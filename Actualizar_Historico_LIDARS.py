import pandas as pd
import numpy as np
from openpyxl import load_workbook # Para trabajar con archivos Excel
import win32com.client as win32
import datetime
from tzlocal import get_localzone
import os
# Importar la librería numpy para manejar arrays y operaciones numéricas
# Importar la librería pandas para manejar DataFrames
# Importar la librería openpyxl para trabajar con archivos Excel


# Definimos nombre del archivo de origen y destino
informe_mtto = r'\\S31889004.services.dekra.com\DKI\EOLICA\MAT REF\BASE DE DATOS ESTACIONES\control_LiDARS\informe_mtto_LIDARS.xlsx'
informe_destino = r'\\S31889004.services.dekra.com\DKI\EOLICA\MAT REF\BASE DE DATOS ESTACIONES\control_LiDARS\Excel_Padre_LIDARS.xlsx'
# Definimos la ruta del archivo que se usará como plantilla para crear nuevos informes historicos
plantilla_historico = r'\\S31889004.services.dekra.com\DKI\EOLICA\MAT REF\BASE DE DATOS ESTACIONES\control_LiDARS\HISTORICO LIDARS\plantilla_historico_LIDARS.xlsx'
# Definimos la ruta de la carpeta donde irán los excels con el histórico de cada LIDAR
carpeta_destino_historico = r'\\S31889004.services.dekra.com\DKI\EOLICA\MAT REF\BASE DE DATOS ESTACIONES\control_LiDARS\HISTORICO LIDARS'
# Definimos nombre de las hojas de origen y destino
hoja_mtto = 'Hoja1' # Nombre de la hoja del archivo de origen
hoja_destino = 'Lidars' #Nombre de la hoja del archivo destino
hoja_plantilla_historico = 'Hoja1' # Nombre de la hoja presente en la plantilla



# Importar las funciones de mantenimiento desde el módulo mantenimiento.py
# Asegúrate de que el archivo mantenimiento.py esté en el mismo directorio o en el PYTHONPATH
from funciones_Mantenimiento_LIDARS import (leer_datos_origen, check_file_closed, actualizar_fecha_destino, actualizar_comentario_destino, enviar_correo, \
                                 actualizar_operarios_destino, actualizar_filtros_destino, actualizar_liquido_destino, actualizar_metanol_destino, \
                                 actualizar_baterias_destino, actualizar_escobilla_destino, actualizar_incidencias_destino, actualizar_sensores_destino, \
                                actualizar_bomba_destino, actualizar_extintor_destino, actualizar_datos_destino, encontrar_fila_historico, leer_correos_origen)





def main():
    """Este script actualiza el archivo destino con los datos del archivo de origen.
    Asegúrate de que las rutas de los archivos y los nombres de las hojas sean correctos."""

    # 0. Comprueba que los archivos a modificar estén cerrados:
    informe_destino_cerrado = check_file_closed(informe_destino)
    if not informe_destino_cerrado:
        print(f"Por favor, cierra el archivo {informe_destino} e inténtalo de nuevo.")
        return  # Salir de la función main si el archivo está abierto

    # 1. Leer datos (usando la versión que también devuelve datos de filtros y escobillas)
    (n_equipo, ubicacion, fecha_ultimomantenimiento, comentario_ultimomantenimiento, operarios_ultimomantenimiento,
    codigo_incidencias) = leer_datos_origen(informe_mtto, hoja_mtto)
    print(codigo_incidencias)

    # 2. Encontrar/Crear la fila en el histórico
    fila_hist = encontrar_fila_historico(n_equipo, ubicacion, carpeta_destino_historico, plantilla_historico, hoja_plantilla_historico, informe_destino, hoja_destino)
    print(f"La fila a actualizar en el histórico es: {fila_hist}")

    # 3. Actualizar todos los datos en los archivos correspondientes
    actualizar_fecha_destino(informe_destino, n_equipo, fecha_ultimomantenimiento, fila_hist, carpeta_destino_historico, ubicacion)
    actualizar_comentario_destino(informe_destino, n_equipo, comentario_ultimomantenimiento, fila_hist, carpeta_destino_historico, ubicacion)
    # --- Llamadas a las nuevas funciones refactorizadas ---
    actualizar_operarios_destino(n_equipo, operarios_ultimomantenimiento, fila_hist, carpeta_destino_historico, ubicacion)
    actualizar_metanol_destino(n_equipo, codigo_incidencias, fila_hist, carpeta_destino_historico, ubicacion)
    actualizar_liquido_destino(n_equipo, codigo_incidencias, fila_hist, carpeta_destino_historico, ubicacion)
    actualizar_filtros_destino(n_equipo, codigo_incidencias, fila_hist, carpeta_destino_historico, ubicacion)
    actualizar_escobilla_destino(n_equipo, codigo_incidencias, fila_hist, carpeta_destino_historico, ubicacion)
    actualizar_bomba_destino(n_equipo, codigo_incidencias, fila_hist, carpeta_destino_historico, ubicacion)
    
    actualizar_extintor_destino(n_equipo, codigo_incidencias, fila_hist, carpeta_destino_historico, ubicacion)
    actualizar_datos_destino(n_equipo, codigo_incidencias, fila_hist, carpeta_destino_historico, ubicacion)
    actualizar_baterias_destino(n_equipo, codigo_incidencias, fila_hist, carpeta_destino_historico, ubicacion)
    actualizar_sensores_destino(informe_destino, n_equipo, codigo_incidencias, fila_hist, carpeta_destino_historico, ubicacion, hoja_destino_name='Lidars')
    actualizar_incidencias_destino(informe_destino, n_equipo, codigo_incidencias, fila_hist, carpeta_destino_historico, ubicacion, hoja_destino_name="Lidars")




    # Anotar la fecha de la Última actualización en el Excel Padre:
    wb_padre = load_workbook(informe_destino)
    hoja_padre = wb_padre[hoja_destino]
    fecha_ult_actualizacion = datetime.datetime.now(get_localzone()).date() # Fecha de hoy en formato date
    fecha_ult_actualizacion = hoja_padre.cell(row=4, column=2, value=fecha_ult_actualizacion)
    fecha_ult_actualizacion.number_format = "DD/MM/YYYY"
    wb_padre.save(informe_destino)


    print("Datos actualizados correctamente.")
    return n_equipo, ubicacion, fecha_ultimomantenimiento, comentario_ultimomantenimiento, operarios_ultimomantenimiento, codigo_incidencias


def main_correo(n_equipo, ubicacion, fecha_ultimomantenimiento, comentario_ultimomantenimiento, operarios_ultimomantenimiento, codigo_incidencias):
    # Leer los correos del archivo de origen
    correos_destino, fecha_envio = leer_correos_origen(informe_mtto, hoja_mtto)
    # Enviar un correo de notificación
    # Escribimos el mensaje con el comentario del último mantenimiento, resaltando los errores a tener en cuneta para el próximo mantenimiento
    operarios_DEKRA, operarios_externos = operarios_ultimomantenimiento
    mensaje = f"""Hola,

    Se ha realizado el mantenimiento del equipo {n_equipo}, en {ubicacion}, a día {fecha_ultimomantenimiento}.
    Siendo los operarios que han realizado el mantenimiento: Por DEKRA: {operarios_DEKRA}, y por parte de WTT: {operarios_externos}.

    Comentario del último mantenimiento:
    {comentario_ultimomantenimiento}

    """

    errores = [incidencia for incidencia in codigo_incidencias if incidencia[0].startswith('Error')]
    if len(errores)>0:
        mensaje += "Por favor, ten en cuenta las siguientes incidencias para el próximo mantenimiento:\n"
        for incidencia in codigo_incidencias:
            mensaje += f"- {incidencia}\n"
    enviar_correo(destinatario=correos_destino, asunto=f"Información para el próximo mantenimientocdel equipo {n_equipo}", \
                  mensaje=mensaje, hora_envio=fecha_envio, cc=None, adjuntos=informe_mtto)


# Ejecutar la función principal
# Esta parte del código se ejecuta cuando el script se corre directamente
# Si se importa como un módulo, no se ejecutará automáticamente, ya que main() no se llamará.
if __name__ == "__main__":
    n_equipo, ubicacion, fecha_ultimomantenimiento, comentario_ultimomantenimiento, operarios_ultimomantenimiento, codigo_incidencias = main()
    main_correo(n_equipo, ubicacion, fecha_ultimomantenimiento, comentario_ultimomantenimiento, operarios_ultimomantenimiento, codigo_incidencias)
    