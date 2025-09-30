# Importar las librerías necesarias
from openpyxl import load_workbook # Para trabajar con archivos Excel
import datetime # import datetime # Para manejar fechas
import sys # Para manejar errores y salir del programa
import pandas as pd # Para manejar DataFrames
import numpy as np # Para manejar arrays y operaciones numéricas
import shutil # Para crear excels
import win32com.client as win32 # Para enviar correos con Outlook
from tzlocal import get_localzone # Para manejar zonas horarias
import os
from copy import copy




# === Funciones principales: Lectura de datos ===
def leer_datos_origen(informe_mtto, hoja_mtto):
    """
    Lee los datos del archivo de origen, y devuelve un DataFrame si y solo si los datos tienen el formsto correcto.
    """
    df = pd.read_excel(informe_mtto, sheet_name=hoja_mtto, engine='openpyxl')

    # Comprobar que el DataFrame tiene al menos una fila y las columnas necesarias
    if df.empty or df.shape[1] < 19:
        raise ValueError("El DataFrame está vacío o no tiene suficientes columnas.")
    
    # El formato de la fecha es especialmente sensible y puede dar problemas, así que lo normalizamos con una función auxiliar:
    def leer_fecha(df, fila, columna=1):
        """
        Extrae la fecha de una celda de un DataFrame y la normaliza a datetime.date.
        """
        valor = df.iloc[fila, columna]

        # Si el valor no es un objeto de fecha nativo, intentamos convertirlo
        if not isinstance(valor, (datetime.datetime, datetime.date, pd.Timestamp)):
            try:
                # Usamos pd.to_datetime, que es más robusto para cadenas de fecha.
                # dayfirst=True asegura que '25/08' se interprete como 25 de agosto.
                valor = pd.to_datetime(valor, dayfirst=True)
            except (ValueError, TypeError) as e:
                raise TypeError(f"El valor de fecha '{valor}' en la fila {fila}, columna {columna} tiene un formato no reconocido. Por favor, asegúrate de que es una fecha válida. Error original: {e}")

        # Normalizamos a datetime.date para que el formato sea consistente
        if isinstance(valor, (datetime.datetime, pd.Timestamp)):
            return valor.date()
        elif isinstance(valor, datetime.date):
            return valor
        else:
            # Este caso es poco probable, pero se mantiene como un control de seguridad
            raise TypeError(f"Tipo inesperado después de la conversión: {type(valor)}")

    n_equipo = df.iloc[0, 0]  # Obtiene valor n equipo LiDAR
    ubicacion = df.iloc[2, 0] # Obtiene valor de la ubicacion de LIDAR
    fecha_ultimomantenimiento = leer_fecha(df, 0, 1)  # Obtiene valor fecha del ultimo mantenimiento
    comentario_ultimomantenimiento = df.iloc[0, 16]  # Obtiene valor comentario del ultimo mantenimiento
    
    # Se genera un código para incidencias (y otros datos) según los valores leídos en el DataFrame:
    codigo_incidencias = []

    # Obtenemos booleano: Se añade liquido limpiaparabrisas?
    adicion_liquido = df.iloc[0, 5]
    assert isinstance(adicion_liquido, (int, np.integer, float, np.floating)), \
        f"Liquido añadido debe ser int o float, no {type(adicion_liquido)}"
    adicion_liquido = False if pd.isna(adicion_liquido) or adicion_liquido == 0 else True
    # Obtiene valor liquido parabrisas del ultimo mantenimiento [Cambio liquido (booleano), Liquido añadido (int), Liquido restante (int)]  
    liquido_ultimomantenimiento = [adicion_liquido , df.iloc[0, 5], df.iloc[0, 6]] 
    codigo_incidencias.append(["Líquido", liquido_ultimomantenimiento])

    # Obtenemos booleano: Se añade metanol?
    adicion_metanol = df.iloc[0, 7]
    assert isinstance(adicion_metanol, (int, np.integer, float, np.floating)), \
        f"Metanol añadido debe ser numérico, no {type(adicion_metanol)}"
    adicion_metanol = False if pd.isna(adicion_metanol) or adicion_metanol == 0 else True
    porcentaje_cartucho1 = df.iloc[2, 7]  # Porcentaje cartucho 1
    porcentaje_cartucho2 = df.iloc[2, 8]  # Porcentaje cartucho 2
    # Obtiene valor metanol del ultimo mantenimiento [Cambio metanol (booleano), Cartuchos añadidos al EFOY (int), Cartuchos añadidos al STOCK (int), \
    # Porcentaje cartucho 1 (int), Porcentaje cartucho 2 (int)]
    metanol_ultimomantenimiento = [adicion_metanol , df.iloc[0, 7], df.iloc[0, 8], porcentaje_cartucho1, porcentaje_cartucho2]
    codigo_incidencias.append(["Metanol", metanol_ultimomantenimiento])

    # Normalizamos los valores booleanos de las columnas que contienen 'SI'/'NO'
    columnas_si_no = [2, 3, 4, 9, 10, 11, 12] # Columnas que contienen 'SI'/'NO'
    for col in columnas_si_no:
        valor = str(df.iloc[0, col]).strip().upper()  # normaliza texto
        if valor == "SI":
            df.iloc[0, col] = True
        elif valor == "NO":
            df.iloc[0, col] = False
        else:
            raise ValueError(f"Valor inesperado en la columna {col}: {valor}. Debe ser 'SI' o 'NO'.")
    

    if df.iloc[0, 2]==False: # EFOY funciona?
        print("Advertencia: EFOY no funciona")
        codigo_incidencias.append("Error_EFOY")
    efoy_online = str(df.iloc[2, 2]).strip().upper()  # EFOY online?
    if efoy_online =="SI": 
        pass
    elif efoy_online =="NO":
        codigo_incidencias.append("Error_EFOY_Offline")
    else:
        print("Advertencia: No se ha especificado si el EFOY está online o no")
        sys.exit(1)

    # Lectura del estado de los filtros:
    n_recambios_filtros = df.iloc[4, 3]  # Número de recambios de filtros
    assert isinstance(n_recambios_filtros, (int, np.integer, float)), f"Número de recambios de filtros debe ser int, no {type(n_recambios_filtros)}"
    if n_recambios_filtros < 0 or None:
        print("Advertencia: Especifique el número de recambios de filtros")
        sys.exit(1)
    if df.iloc[0, 3]==True: # Filtros de Lidar sustituidos?
        print("Advertencia: Filtros de Lidar se han sustituido")
        filtros_desechados = df.iloc[2, 3]  # Nº de Filtros desechados
        assert isinstance(filtros_desechados, (int, np.integer, float)), f"Número de recambios de filtros debe ser int, no {type(n_recambios_filtros)}"
        if filtros_desechados==(0 or None): 
            codigo_incidencias.append(["Error_Filtro_Sustituido", [filtros_desechados, n_recambios_filtros]])
        elif filtros_desechados >0:
            codigo_incidencias.append(["Error_Filtro_Desechado", [filtros_desechados, n_recambios_filtros]])
        elif filtros_desechados <0:
            print("Advertencia: El número de filtros desechados no puede ser negativo")
            sys.exit(1)
    else:
        filtros_desechados = 0
        codigo_incidencias.append(["Filtro", [filtros_desechados, n_recambios_filtros]])
    print(f"Nº de filtros desechados, nº recambios de filtros: {[filtros_desechados, n_recambios_filtros]}")

    
    # Lectura del estado de la escobilla:
    n_recambios_escobillas = df.iloc[2, 4]  # Número de recambios de escobillas
    assert isinstance(n_recambios_escobillas, (int, np.integer, float)), f"Número de recambios de escobillas debe ser int, no {type(n_recambios_escobillas)}"
    if n_recambios_escobillas < 0:
        print("Advertencia: Número de recambios de escobillas no puede ser negativo")
        sys.exit(1)
    if df.iloc[0, 4]==False: # Escobilla correcta?
        print("Advertencia: Escobilla no esta bien")
        codigo_incidencias.append(["Error_Escobilla", n_recambios_escobillas])
    else:
        codigo_incidencias.append(["Escobilla", n_recambios_escobillas])


    # Lectura del estado de las baterías y sus detalles
    if df.iloc[0, 9] == True: # Se cambian las baterías? 
        print("Advertencia: Se han cambiado las baterías")
        # Leer el SOH y los números de serie desde las celdas debajo
        soh_baterias = df.iloc[1:, 9].dropna().tolist()
        # Agregar la incidencia con los detalles
        codigo_incidencias.append(["Error_Baterias", soh_baterias])
    else:
        soh_baterias = df.iloc[1:, 9].dropna().tolist()
        if soh_baterias:
            print("No se cambian las baterías, pero se han registrado sus estados.")
            codigo_incidencias.append(["Estado_Baterias", soh_baterias])
        else:
            print("No se ha registrado estados de baterías")


    #Lectura del estado de la bomba de agua
    n_recambios_bomba = df.iloc[2, 10]  # Número de recambios de bomba de agua
    assert isinstance(n_recambios_bomba, (int, np.integer, float)), f"Número de recambios de bomba de agua debe ser int, no {type(n_recambios_bomba)}"
    if n_recambios_bomba < 0:
        print("Advertencia: Número de recambios de bomba de agua no puede ser negativo")
        sys.exit(1)
    if df.iloc[0, 10]==False: # Bomba de agua funciona?
        print("Advertencia: Bomba de agua no funciona")
        codigo_incidencias.append(["Error_Bomba", n_recambios_bomba])
    else:
        codigo_incidencias.append(["Bomba", n_recambios_bomba])

    # Lectura del estado del extintor de incendios
    fecha_caducidad_extintor = leer_fecha(df, 2, 11)  # Fecha de caducidad del extintor
    meses_para_caducidad = (fecha_caducidad_extintor.year - fecha_ultimomantenimiento.year) * 12 + \
                             (fecha_caducidad_extintor.month - fecha_ultimomantenimiento.month)
    if df.iloc[0, 11]==False: # Hay extintor?
        print("Advertencia: No hay extintor de incendios")
    elif meses_para_caducidad > 3:
        # Se añade al final porque es un dato importante para el historico, aunque no sea una incidencia
        codigo_incidencias.append(["Fecha_Extintor", fecha_caducidad_extintor])
    elif 0 <= meses_para_caducidad <= 3:
        print("Advertencia: Extintor de incendios está próximo a caducar")
        codigo_incidencias.append(["Error_Extintor", fecha_caducidad_extintor])
    else: # meses_para_caducidad < 0
        print("Advertencia: Extintor de incendios ha caducado")
        codigo_incidencias.append(["Error_Extintor_Caducado", fecha_caducidad_extintor])
    
    
    
    # Lectura de la descarga de datos
    fecha_inicio_datos = leer_fecha(df, 2, 12)  # Fecha inicio de datos
    fecha_fin_datos = leer_fecha(df, 4, 12)    
    if df.iloc[0, 12]==False: # Descarga de datos?
        print("Advertencia: No se han descargado los datos")
        codigo_incidencias.append(["Error_DescargaDatos", []])
    else:
        # Se añade el período de los datos en el historico en caso que se hayan descargado
        codigo_incidencias.append(["Fechas_Datos", [fecha_inicio_datos, fecha_fin_datos]])

    # Lectura de los sensores y sus detalles
    sensores_cambiados_ultimomantenimiento = []
    if df.iloc[0, 13]: # Sensores cambiados?
        # Contar cuantas columnas de sensores cambiados hay:
        sensor_nombres = df.iloc[0:, 13].dropna().tolist()
        nserie_cambio = df.iloc[0:, 14].dropna().tolist()
        nserie_recambio = df.iloc[0:, 15].dropna().tolist()
        
        if len(sensor_nombres) > 0:
            print("Se han cambiado los siguientes sensores:", sensor_nombres)
            sensores_cambiados_ultimomantenimiento = list(zip(sensor_nombres, nserie_cambio, nserie_recambio))
            codigo_incidencias.append(["Error_Sensores", sensores_cambiados_ultimomantenimiento])
        else:
            print("No se han cambiado sensores.")
    else:
        print("No se han cambiado sensores.")


    # Lectura de los operarios
    operarios_DEKRA = df.iloc[6:, 0].dropna().tolist()  # Operarios DEKRA
    operarios_externos = df.iloc[6:, 1].dropna().tolist()  # Operarios externos
    for operario in operarios_DEKRA + operarios_externos:
        assert isinstance(operario, str), f"Operario debe ser str, no {type(operario)}"
    operarios_ultimomantenimiento = [operarios_DEKRA, operarios_externos]
    if not operarios_DEKRA and not operarios_externos:
        print("Advertencia: No se han especificado los operarios que realizaron el mantenimiento")
        sys.exit(1)

    # Si no hay incidencias, se añade un mensaje de que todo esta correcto:
    if not codigo_incidencias:
        print("No hay incidencias, todo correcto.")

    # Comprueba el formato de los datos leídos:
    assert isinstance(n_equipo, (str, int)), f"n_equipo debe ser str o int, no {type(n_equipo)}"
    assert isinstance(comentario_ultimomantenimiento, str), f"comentario_ultimomantenimiento debe ser str, no {type(comentario_ultimomantenimiento)}"
    assert isinstance(ubicacion, str), f"ubicacion debe ser str, no {type(ubicacion)}"
    assert isinstance(liquido_ultimomantenimiento[0], (bool, int)), f"Cambio liquido debe ser bool o int, no {type(liquido_ultimomantenimiento[0])}"
    assert isinstance(liquido_ultimomantenimiento[1],(int, np.integer, float, np.floating)), f"Liquido añadido debe ser int o float, no {type(liquido_ultimomantenimiento[1])}"
    assert isinstance(liquido_ultimomantenimiento[2],(int, np.integer, float, np.floating)), f"Liquido restante debe ser int o float, no {type(liquido_ultimomantenimiento[2])}"
    assert isinstance(metanol_ultimomantenimiento[0], (bool, int)), f"Cambio metanol debe ser bool o int, no {type(metanol_ultimomantenimiento[0])}"
    assert isinstance(metanol_ultimomantenimiento[1],(int, np.integer, float, np.floating)), f"Cartuchos añadidos al EFOY debe ser int o float, no {type(metanol_ultimomantenimiento[1])}"
    assert isinstance(metanol_ultimomantenimiento[2],(int, np.integer, float, np.floating)), f"Cartuchos añadidos al STOCK debe ser int o float, no {type(metanol_ultimomantenimiento[2])}"
    assert isinstance(metanol_ultimomantenimiento[3],(int, np.integer, float, np.floating)), f"Porcentaje cartucho 1 debe ser int o float, no {type(metanol_ultimomantenimiento[3])}"
    assert isinstance(metanol_ultimomantenimiento[4],(int, np.integer, float, np.floating)), f"Porcentaje cartucho 2 debe ser int o float, no {type(metanol_ultimomantenimiento[4])}"
    assert isinstance(df.iloc[0, 2], bool), f"EFOY funciona debe ser SI o NO, no {type(df.iloc[0, 2])}"
    assert isinstance(df.iloc[0, 3], bool), f"Filtros de Lidar sustituidos debe ser SI o NO, no {type(df.iloc[0, 3])}"
    assert isinstance(df.iloc[0, 4], bool), f"Escobilla limpia debe ser SI o NO, no {type(df.iloc[0, 4])}"
    assert isinstance(df.iloc[0, 9], bool), f"Se cambian las baterías? debe ser SI o NO, no {type(df.iloc[0, 9])}"
    assert isinstance(df.iloc[0, 10], bool), f"Bomba de agua funciona debe ser SI o NO, no {type(df.iloc[0, 10])}"
    assert isinstance(df.iloc[0, 11], bool), f"Extintor de incendios revisado debe ser SI o NO, no {type(df.iloc[0, 11])}"
    assert isinstance(df.iloc[0, 12], bool), f"Descarga de datos debe ser SI o NO, no {type(df.iloc[0, 12])}"
    
    # He eliminado las variables `sensores_cambiados_ultimomantenimiento` y `estado_baterias` del return
    # ya que ahora están contenidas dentro de `codigo_incidencias`.
    print("Datos leídos correctamente del archivo de origen, para el número de equipo:", n_equipo)
    return n_equipo, ubicacion, fecha_ultimomantenimiento, comentario_ultimomantenimiento, operarios_ultimomantenimiento,\
           codigo_incidencias






def leer_correos_origen(informe_mtto, hoja_mtto):
    """
    Lee del archivo de origen:
    - Columna 18 (índice 17): fecha y hora en formato dd-mm-yy hh:mm.
    - Columna 19 (índice 18): lista de correos destino.

    Devuelve:
        (list[str], datetime.datetime)
    """

    df = pd.read_excel(informe_mtto, sheet_name=hoja_mtto, engine='openpyxl')

    # Comprobar que el DataFrame tiene al menos una fila y las columnas necesarias
    if df.empty or df.shape[1] < 19:
        raise ValueError("El DataFrame está vacío o no tiene suficientes columnas.")

    # === 1. Fecha y hora ===
    fecha_envio = df.iloc[0, 17]

    # Normalizar a datetime
    if isinstance(fecha_envio, datetime.date) and not isinstance(fecha_envio, datetime.datetime):
        # convertir date -> datetime con hora 00:00
        fecha_envio = datetime.datetime.combine(fecha_envio, datetime.time.min)

    elif isinstance(fecha_envio, pd.Timestamp):
        fecha_envio = fecha_envio.to_pydatetime()

    elif isinstance(fecha_envio, str):
        # intentar parsear string a datetime
        formatos_posibles = ["%d-%m-%y %H:%M", "%d/%m/%Y %H:%M", "%d/%m/%y %H:%M"]
        for fmt in formatos_posibles:
            try:
                fecha_envio = datetime.datetime.strptime(fecha_envio.strip(), fmt)
                break
            except ValueError:
                continue
        else:
            raise ValueError(
                f"No se pudo interpretar la fecha_envio '{fecha_envio}' con formatos conocidos."
            )

    if not isinstance(fecha_envio, datetime.datetime):
        raise TypeError(f"fecha_envio debe ser datetime, no {type(fecha_envio)}")

    # === 2. Correos destino ===
    correos_destino = df.iloc[:, 18].dropna().tolist()

    if not correos_destino:
        raise ValueError("No se encontraron correos destino en la columna 18.")

    for correo in correos_destino:
        assert isinstance(correo, str), f"correo debe ser str, no {type(correo)}"
        assert "@" in correo and "." in correo, f"correo {correo} no es válido."

    print("Correos y fecha/hora leídos correctamente del archivo de origen.")

    return correos_destino, fecha_envio





# === Helpers ===
def check_file_closed(path: str) -> bool:
    """Return True si el archivo se puede abrir (i.e., no está bloqueado por Excel)."""
    try:
        # Intenta abrir el archivo en modo append
        with open(path, "a+"):
            pass
        return True
    except PermissionError:
        return False
    except Exception as e:
        print(f"Error inesperado al verificar el archivo {path}: {e}")
        return False



def encontrar_fila_destino(wb, n_equipo, hoja_destino_name='Lidar Windcube'):
    """
    Localiza en la hoja destino la fila correspondiente a un equipo (n_equipo).
    Busca en la primera columna desde la fila 3 hacia abajo.
    """
    hoja_destino = wb[hoja_destino_name]
    for idx, (val,) in enumerate(hoja_destino.iter_rows(min_row=3, max_col=1, values_only=True), start=3):
        if val == n_equipo:
            return idx
    raise ValueError(f"No se encontró el equipo {n_equipo} en '{hoja_destino_name}'")


def encontrar_fila_historico(n_equipo, ubicacion, carpeta_destino_historico, plantilla_historico, hoja_plantilla_historico, informe_destino, hoja_destino_name):
    """
    Gestiona el archivo histórico de mantenimientos para un equipo dado.
    
    - Si no existe "n_equipo.xlsx", lo crea a partir de la plantilla y añade una fila en el Excel Padre
    - Si existe:
        * Si existe hoja con nombre "ubicacion", trabaja allí añadiendo fila nueva.
        * Si no existe hoja con nombre "ubicacion", pregunta si se está cambiando de ubicación.
          En caso afirmativo, crea la hoja copiando la estructura de la plantilla.
    
    Devuelve el índice de la fila donde se debe escribir el nuevo mantenimiento.

    ¡¡¡ En la plantilla, la primera fila editable es la 4 !!!
    """

    # --- función auxiliar 1: obtener siguiente fila ---
    def get_next_row(ws, col=1):
        """Devuelve la siguiente fila libre en la columna `col` de la hoja `ws`."""
        last_row = 0
        for row in range(ws.max_row, 0, -1):
            if ws.cell(row=row, column=col).value not in (None, ""):
                last_row = row
                break
        return last_row + 1

    # --- función auxiliar 2: copiar celda a celda ---
    def copy_sheet_contents(ws_source, ws_target):
        """
        Copia valores, estilos, alturas y anchos de columnas de ws_source a ws_target.
        Nota: gráficos, imágenes y macros no se copian con openpyxl.
        """
        # Copiar celdas
        for row in ws_source.iter_rows():
            for cell in row:
                new_cell = ws_target.cell(row=cell.row, column=cell.col_idx, value=cell.value)
                if cell.has_style:
                    new_cell.font = copy(cell.font)
                    new_cell.border = copy(cell.border)
                    new_cell.fill = copy(cell.fill)
                    new_cell.number_format = copy(cell.number_format)
                    new_cell.protection = copy(cell.protection)
                    new_cell.alignment = copy(cell.alignment)

        # Copiar anchos de columnas
        for col_letter, dim in ws_source.column_dimensions.items():
            ws_target.column_dimensions[col_letter].width = dim.width

        # Copiar alturas de filas
        for row_idx, dim in ws_source.row_dimensions.items():
            ws_target.row_dimensions[row_idx].height = dim.height
    # -------------------------------------------------


    # Nombre del excel histórico
    excel_historico = os.path.join(carpeta_destino_historico, f"{n_equipo}.xlsx")



    # Caso 1: No existe el Excel → nuevo LIDAR
    if not os.path.exists(excel_historico):
        while True:
            nuevo_LIDAR = input(f"No existe registro histórico del LIDAR {n_equipo}. ¿Estás registrando un LIDAR nuevo? (s/n): ").strip().lower()
            if nuevo_LIDAR == "s":
                break
            elif nuevo_LIDAR == "n":
                raise ValueError(f"Operación cancelada. Verifica que el número de equipo es {n_equipo}.")
            else:
                print("Respuesta inválida. Introduce 's' o 'n'.")

        shutil.copy(plantilla_historico, excel_historico)
        wb_historico = load_workbook(excel_historico)

        # Renombramos la hoja de la plantilla
        if hoja_plantilla_historico in wb_historico.sheetnames:
            ws = wb_historico[hoja_plantilla_historico]
            ws.title = ubicacion
        else:
            raise ValueError(f"La plantilla no contiene la hoja base '{hoja_plantilla_historico}'.")

        # Escribimos información inicial
        ws.cell(row=5, column=1, value=n_equipo)
        ws.cell(row=3, column=2, value=ubicacion)
        wb_historico.save(excel_historico)

        # Añadimos una fila en el Excel padre
        wb_padre = load_workbook(informe_destino)
        hoja_padre = wb_padre[hoja_destino_name]
        next_row_padre = get_next_row(hoja_padre, 1)
        hoja_padre.cell(row=next_row_padre, column=1, value=n_equipo)
        hoja_padre.cell(row=next_row_padre, column=2, value=ubicacion)
        hoja_padre.cell(row=next_row_padre, column=5, value=datetime.date(1900, 1, 1)) # Fecha penúltimo mantenimiento (inicial)
        country = input(f"Introduce el país donde se encuentra el LIDAR {n_equipo}: ").strip()
        hoja_padre.cell(row=next_row_padre, column=3, value=country)
        client = input(f"Introduce el cliente asociado al LIDAR {n_equipo}: ").strip()
        hoja_padre.cell(row=next_row_padre, column=4, value=client)
        wb_padre.save(informe_destino)

        return 5

    # Caso 2: Ya existe el Excel
    else:
        check_file_closed(excel_historico) # Verificar que el archivo no esté abierto
        excel_historico_cerrado = check_file_closed(excel_historico)
        if not excel_historico_cerrado:
            raise ValueError(f"Por favor, cierra el archivo {excel_historico} e inténtalo de nuevo.")
        

        wb_historico = load_workbook(excel_historico)

        # Caso 2a: existe la hoja (mantenimiento en ubicación conocida)
        if ubicacion in wb_historico.sheetnames:
            ws = wb_historico[ubicacion]
            next_row = get_next_row(ws, 1)
            ws.cell(row=next_row, column=1, value=n_equipo)
            wb_historico.save(excel_historico)
            return next_row

        # Caso 2b: no existe la hoja → preguntar si se crea
        else:
            while True:
                cambio_ubicacion = input(f"No existe hoja '{ubicacion}' en {excel_historico}. ¿Crear nueva ubicación? (s/n): ").strip().lower()

                if cambio_ubicacion == "s":
                    wb_plantilla = load_workbook(plantilla_historico)
                    if hoja_plantilla_historico not in wb_plantilla.sheetnames:
                        raise ValueError(f"La plantilla '{plantilla_historico}' no contiene la hoja '{hoja_plantilla_historico}'.")

                    # Crear hoja vacía y copiar estructura
                    ws_template = wb_plantilla[hoja_plantilla_historico]
                    ws_historico = wb_historico.create_sheet(title=ubicacion)
                    copy_sheet_contents(ws_template, ws_historico)

                    # Escribir datos iniciales
                    ws_historico.cell(row=5, column=1, value=n_equipo)
                    ws_historico.cell(row=3, column=2, value=ubicacion)

                    wb_historico.save(excel_historico)
                    return 5

                elif cambio_ubicacion == "n":
                    raise ValueError(f"Operación cancelada. Verifique la ubicación {ubicacion}.")
                else:
                    print("Respuesta inválida. Introduce 's' o 'n'.")


def leer_int(ws, fila, col):
            """Lee un valor entero de una celda, devolviendo 0 si está vacía o no es un entero."""
            valor = ws.cell(row=fila, column=col).value
            try:
                return int(valor)
            except (TypeError, ValueError):
                return 0  # si está vacío devuelves 0, si no y no es int, error





# === Funciones principales: Actulización de datos ===

def actualizar_fecha_destino(informe_destino, n_equipo, fecha_ultimomantenimiento, fila_hist,
                             carpeta_destino_historico, ubicacion,
                             hoja_destino_name="Lidar Windcube"):
    """
    Actualiza la fecha del último mantenimiento en:
    1. El Excel padre (hoja 'Lidar Windcube').
    2. La fila correspondiente del histórico de ese LIDAR (columna 2).

    Parámetros:
        informe_destino (str): Ruta del Excel padre.
        n_equipo (str|int): Identificador del LIDAR.
        fecha_ultimomantenimiento (datetime): Fecha a registrar.
        fila_hist (int): Número de fila en el histórico donde se está registrando este mantenimiento.
        carpeta_destino_historico (str): Carpeta donde están los excels de histórico.
        ubicacion (str): Nombre de la hoja en el histórico (ubicación del LIDAR).
        hoja_destino_name (str): Nombre de la hoja del Excel padre (default: 'Lidar Windcube').
    """
    wb_padre = None
    wb_hist = None

    try:
        # === 1. Actualizar en Excel padre ===
        wb_padre = load_workbook(informe_destino)
        hoja_padre = wb_padre[hoja_destino_name]

        # Función auxiliar para encontrar la fila del equipo
        def encontrar_fila_destino(wb, n_equipo, hoja_name):
            ws = wb[hoja_name]
            for i, row in enumerate(ws.iter_rows(min_row=3, values_only=True), 3):
                if row[0] == n_equipo:
                    return i
            return None

        fila_destino = encontrar_fila_destino(wb_padre, n_equipo, hoja_destino_name)
        if not fila_destino:
            print(f"Advertencia: No se encontró el equipo {n_equipo} en el archivo padre.")
            sys.exit(1)
        
        fecha_penultima = hoja_padre.cell(row=fila_destino, column=5).value
        
        # Lógica de validación estricta
        if not fecha_penultima:
            print(f"Error: La celda de fecha para el equipo {n_equipo} en el archivo padre está vacía. No se puede continuar.")
            sys.exit(1)
        
        if isinstance(fecha_penultima, datetime.datetime):
            fecha_penultima_norm = fecha_penultima.date()
        elif isinstance(fecha_penultima, datetime.date):
            fecha_penultima_norm = fecha_penultima
        else:
            fecha_penultima_norm = None

        if fecha_penultima_norm == fecha_ultimomantenimiento:
            print(f"La fecha {fecha_penultima_norm.strftime('%d/%m/%Y')} ya está registrada para el equipo {n_equipo}. No se actualiza.")
            print("Si deseas relizar la actualización, elimina la fila creada en el histórico y la casilla en el excel padre y vuelve a ejecutar el script.")

            # === 2. Actualizar en Excel histórico, borrando la fila creada
            excel_historico = os.path.join(carpeta_destino_historico, f"{n_equipo}.xlsx")
            wb_hist = load_workbook(excel_historico)
            ws_hist = wb_hist[ubicacion]
            ws_hist.delete_rows(fila_hist, 1)
            wb_hist.save(excel_historico)
            sys.exit(1)
        
        # Escribir nueva fecha en hoja padre
        celda = hoja_padre.cell(row=fila_destino, column=5, value=fecha_ultimomantenimiento)
        celda.number_format = "DD/MM/YYYY"
        wb_padre.save(informe_destino)
        
        # === 2. Actualizar en Excel histórico ===
        excel_historico = os.path.join(carpeta_destino_historico, f"{n_equipo}.xlsx")
        wb_hist = load_workbook(excel_historico)
        ws_hist = wb_hist[ubicacion]

        # Escribir la fecha en la columna 2 de la fila indicada
        celda_hist = ws_hist.cell(row=fila_hist, column=2, value=fecha_ultimomantenimiento)
        celda_hist.number_format = "DD/MM/YYYY"
        wb_hist.save(excel_historico)
        
        fecha_str = fecha_ultimomantenimiento.strftime("%d/%m/%Y")
        print(f"✔ Fecha {fecha_str} registrada correctamente para {n_equipo} en {ubicacion}, fila {fila_hist}.")

    except Exception as e:
        print(f"Error al actualizar la fecha: {e}")
        sys.exit(1)
    finally:
        if wb_padre:
            wb_padre.close()
        if wb_hist:
            wb_hist.close()

def actualizar_comentario_destino(informe_destino, n_equipo, comentario_ultimomantenimiento, fila_hist,
                             carpeta_destino_historico, ubicacion, hoja_destino_name="Lidar Windcube"):
    """
    Actualiza el archivo destino con el comentario del último mantenimiento:
    - Sustituye el comentario en la hoja destino.
    - Añade el comentario nuevo al historico
    """

    # === 1. Actualizar en Excel padre ===
    wb_padre = load_workbook(informe_destino)
    hoja_padre = wb_padre[hoja_destino_name]
    fila_destino = encontrar_fila_destino(wb_padre, n_equipo, hoja_destino_name)


    # Escribir nuevo comentario en hoja padre
    hoja_padre.cell(row=fila_destino, column=8, value=comentario_ultimomantenimiento)
    wb_padre.save(informe_destino)
    wb_padre.close()

    # === 2. Actualizar en Excel histórico ===
    excel_historico = os.path.join(carpeta_destino_historico, f"{n_equipo}.xlsx")
    wb_hist = load_workbook(excel_historico)
    ws_hist = wb_hist[ubicacion]

    # Escribir la fecha en la columna 3 de la fila indicada
    ws_hist.cell(row=fila_hist, column=3, value=comentario_ultimomantenimiento)
    wb_hist.save(excel_historico)
    wb_hist.close()

    print(f"Comentario registrado correctamente")


def actualizar_operarios_destino(n_equipo, operarios_ultimomantenimiento, fila_hist,
                                 carpeta_destino_historico, ubicacion):
    """
    Actualiza el archivo destino con los operarios del último mantenimiento:
    - Añade los operarios en el historico
    """
    # --- 1. Abrir y actualizar el archivo histórico ---
    excel_historico = os.path.join(carpeta_destino_historico, f"{n_equipo}.xlsx")
    wb_hist = load_workbook(excel_historico)
    ws_hist = wb_hist[ubicacion]
    operarios_DEKRA, operarios_externos = operarios_ultimomantenimiento
    # Convertir listas de operarios a cadenas separadas por comas   
    operarios_DEKRA_str = ", ".join(operarios_DEKRA) if operarios_DEKRA else "N/A"
    operarios_externos_str = ", ".join(operarios_externos) if operarios_externos else "N/A"
    operarios_str = f"DEKRA: {operarios_DEKRA_str}; Externos: {operarios_externos_str}"
    # Escribir en la columna 16, guardar y cerrar el archivo:
    ws_hist.cell(row=fila_hist, column=16, value=operarios_str)
    wb_hist.save(excel_historico)
    wb_hist.close()


def actualizar_metanol_destino(n_equipo, codigo_incidencias, fila_hist, carpeta_destino_historico, ubicacion):
    """
    Actualiza la información de metanol:
    1. En el Excel histórico del LIDAR (n_equipo.xlsx, hoja = ubicacion):
        - Columna 4: metanol en stock tras el mantenimiento y porcentaje de los cartuchos.
        - Columna 5: total de metanol usado por el LiDAR.

    Parámetros:
        n_equipo (str|int): Identificador del LIDAR.
        Dentro del codigo_incidencias:
        - metanol_ultimomantenimiento (tuple): (cambio_metanol, añadido_EFOY, añadido_stock, porcentaje_cartucho1, porcentaje_cartucho2).
        fila_hist (int): Fila en el histórico donde escribir.
        carpeta_destino_historico (str): Carpeta de históricos individuales.
        ubicacion (str): Nombre de la hoja del histórico.
    """
    # --- 1. Abrir y actualizar el archivo histórico ---
    excel_historico = os.path.join(carpeta_destino_historico, f"{n_equipo}.xlsx")
    wb_hist = load_workbook(excel_historico)
    ws_hist = wb_hist[ubicacion]

    # Desempaquetar valores de metanol
    metanol_ultimomantenimiento = None
    for incidencia in codigo_incidencias:
        if incidencia[0] == "Metanol":
            metanol_ultimomantenimiento = incidencia[1]
            break
    cambio_metanol, metanol_anadido_EFOY, metanol_anadido_stock, porcentaje_cartucho1, porcentaje_cartucho2 = metanol_ultimomantenimiento

    if cambio_metanol:
        # Stock actual previo
        stock_prev = leer_int(ws_hist, fila_hist - 1, 4)
        usado_prev = leer_int(ws_hist, fila_hist - 1, 5)

        # Nuevo stock = stock previo + lo añadido a stock - lo consumido por EFOY
        stock_nuevo = stock_prev + metanol_anadido_stock - metanol_anadido_EFOY
        
        # Combinar el stock con los porcentajes en una única cadena de texto
        valor_celda = f"{stock_nuevo} ({porcentaje_cartucho1} %, {porcentaje_cartucho2} %)"
        ws_hist.cell(row=fila_hist, column=4, value=valor_celda)

        # Total usado = usado previo + lo consumido por EFOY
        usado_nuevo = usado_prev + metanol_anadido_EFOY
        ws_hist.cell(row=fila_hist, column=5, value=usado_nuevo)

        print(f"✔ Metanol actualizado: stock={stock_nuevo}, usado_total={usado_nuevo}")
    else:
        print("ℹ En este mantenimiento no se cambió el metanol")

    wb_hist.save(excel_historico)
    wb_hist.close()


def actualizar_liquido_destino(n_equipo, codigo_incidencias, fila_hist, carpeta_destino_historico, ubicacion):
    """
    Actualiza en el Excel histórico del LIDAR (n_equipo.xlsx, hoja = ubicacion) 
    la información relacionada con el líquido limpiaparabrisas:

    - Columna 6: cantidad de líquido restante tras el mantenimiento.
    - Columna 7: cantidad total de líquido usado en el mantenimiento

    Parámetros:
        n_equipo (str|int): Identificador del LIDAR.
        Dentro del codigo_incidencias:
        - liquido_ultimomantenimiento (tuple): (adicion_liquido, litros_anadidos, litros_restantes).
        fila_hist (int): Fila en el histórico donde escribir.
        carpeta_destino_historico (str): Carpeta de históricos individuales.
        ubicacion (str): Nombre de la hoja del histórico.
    """

    # Abrir el archivo histórico
    excel_historico = os.path.join(carpeta_destino_historico, f"{n_equipo}.xlsx")
    wb_hist = load_workbook(excel_historico)
    ws_hist = wb_hist[ubicacion]

    # Desempaquetar valores de líquido
    liquido_ultimomantenimiento = None
    for incidencia in codigo_incidencias:
        if incidencia[0] == "Líquido":
            liquido_ultimomantenimiento = incidencia[1]
            break
    adicion_liquido, liquido_anadido, liquido_restante = liquido_ultimomantenimiento

    if adicion_liquido:  
        # Escrivimos la cantidad de liquido restante después del mantenimiento
        ws_hist.cell(row=fila_hist, column=6, value=liquido_restante)
        # Y l cantidad de líquido añadido
        ws_hist.cell(row=fila_hist, column=7, value=liquido_anadido)
        print(f"✔ Líquido actualizado: restante={liquido_restante}, usado_total={liquido_anadido}")
    else:  
        # Si no se añadió líquido, solo registrar el restante informado
        ws_hist.cell(row=fila_hist, column=6, value=liquido_restante)
        print("ℹ En este mantenimiento no se añadió líquido limpiaparabrisas")

    wb_hist.save(excel_historico)
    wb_hist.close()



def actualizar_filtros_destino(n_equipo, codigo_incidencias, fila_hist, carpeta_destino_historico, ubicacion):
    """
    Actualiza el contador de filtros restantes en el histórico.
    - Columna 8: Recambios de Filtro restantes.
    """
    # Buscar el valor de "recambios restantes" en la lista de incidencias
    recambios_filtros_restantes = None
    for incidencia in codigo_incidencias:
        if incidencia[0] in ("Error_Filtro_Sustituido", "Error_Filtro_Desechado", "Filtro"):
            filtros_desechados = incidencia[1][0]
            recambios_filtros_restantes = incidencia[1][1]
            break
    
    if recambios_filtros_restantes is None:
        raise ValueError("Error: No se ha especificado el stock restante para los filtros en la lista de incidencias.")
    
    excel_historico = os.path.join(carpeta_destino_historico, f"{n_equipo}.xlsx")
    wb_hist = load_workbook(excel_historico)
    ws_hist = wb_hist[ubicacion]

    # Actualizar filtros - DIRECTAMENTE PEGANDO EL VALOR EN LAS COLUMNAS CORRESPONDIENTES
    ws_hist.cell(row=fila_hist, column=8, value=recambios_filtros_restantes)
    ws_hist.cell(row=fila_hist, column=9, value=filtros_desechados)
    print(f"✔ Filtros actualizados: [Desechados={filtros_desechados}, Recambios disponibles={recambios_filtros_restantes}].")

    wb_hist.save(excel_historico)
    wb_hist.close()


def actualizar_escobilla_destino(n_equipo, codigo_incidencias, fila_hist, carpeta_destino_historico, ubicacion):
    """
    Actualiza el contador de escobillas restantes en el histórico.
    - Columna 10: Recambio de Escobillas restantes.
    """
    recambios_escobillas_restantes = None
    for incidencia in codigo_incidencias:
        if incidencia[0] in ("Error_Escobilla", "Escobilla"):
            recambios_escobillas_restantes = incidencia[1]
            break


    if recambios_escobillas_restantes is None:
        raise ValueError("Error: No se ha especificado el stock restante para la escobilla en la lista de incidencias.")

    excel_historico = os.path.join(carpeta_destino_historico, f"{n_equipo}.xlsx")
    wb_hist = load_workbook(excel_historico)
    ws_hist = wb_hist[ubicacion]

    # Actualizar recambios restantes - DIRECTAMENTE PEGANDO EL VALOR
    ws_hist.cell(row=fila_hist, column=10, value=recambios_escobillas_restantes)
    print(f"✔ Escobilla actualizada: recambios restantes es ahora {recambios_escobillas_restantes}.")

    wb_hist.save(excel_historico)
    wb_hist.close()


def actualizar_bomba_destino(n_equipo, codigo_incidencias, fila_hist, carpeta_destino_historico, ubicacion):
    """
    Actualiza el contador de bombas de agua restantes en el histórico.
    - Columna 15: Recambios de Bomba de Agua restantes.
    """
    recambios_bomba_restantes = None
    for incidencia in codigo_incidencias:
        if incidencia[0] in ("Error_Bomba", "Bomba"):
            recambios_bomba_restantes = incidencia[1]
            break

    if recambios_bomba_restantes is None:
        raise ValueError("Error: No se ha especificado el stock restante para la bomba en la lista de incidencias.")

    excel_historico = os.path.join(carpeta_destino_historico, f"{n_equipo}.xlsx")
    wb_hist = load_workbook(excel_historico)
    ws_hist = wb_hist[ubicacion]

    # Actualizar recambios restantes - DIRECTAMENTE PEGANDO EL VALOR
    ws_hist.cell(row=fila_hist, column=15, value=recambios_bomba_restantes)
    print(f"✔ Bomba de agua actualizada: recambios restantes es ahora {recambios_bomba_restantes}.")

    wb_hist.save(excel_historico)
    wb_hist.close()



 # REMODELAJE:


def actualizar_extintor_destino(n_equipo, codigo_incidencias, fila_hist, carpeta_destino_historico, ubicacion):
    """
    Actualiza la fecha de caducidad del extintor en el histórico.
    - Columna 13: Fecha de caducidad del extintor.
    """

    fecha_caducidad=None

    for incidencia in codigo_incidencias:
        if incidencia[0] in ("Error_Extintor", "Error_Extintor_Caducado", "Fecha_Extintor"):
            fecha_caducidad = incidencia[1]
            break
    if fecha_caducidad is None:
        raise ValueError("Error: No se ha especificado la fecha de caducidad del extintor.")

    excel_historico = os.path.join(carpeta_destino_historico, f"{n_equipo}.xlsx")
    wb_hist = load_workbook(excel_historico)
    ws_hist = wb_hist[ubicacion]
    
    # Formatear la fecha como cadena (asumiendo formato DD-MM-YYYY)
    if isinstance(fecha_caducidad, datetime.datetime):
        fecha_str = fecha_caducidad.strftime("%d-%m-%Y")
    else:
        fecha_str = str(fecha_caducidad)
    

    ws_hist.cell(row=fila_hist, column=13, value=fecha_str)
    print(f"✔ Extintor actualizado: fecha de caducidad es {fecha_str}.")

    wb_hist.save(excel_historico)
    wb_hist.close()


def actualizar_datos_destino(n_equipo, codigo_incidencias, fila_hist, carpeta_destino_historico, ubicacion):
    """
    Actualiza el intervalo de fechas de la descarga de datos en el histórico.
    - Columna 14: Fecha de la descarga de datos.
    """
    fecha_desde=None
    fecha_hasta=None

    excel_historico = os.path.join(carpeta_destino_historico, f"{n_equipo}.xlsx")
    wb_hist = load_workbook(excel_historico)
    ws_hist = wb_hist[ubicacion]
    for incidencia in codigo_incidencias:

        if incidencia[0] in ("Fechas_Datos"):
            fecha_desde, fecha_hasta = incidencia[1]

            # Formatear las fechas como cadenas y combinarlas
            if isinstance(fecha_desde, datetime.datetime):
                fecha_desde_str = fecha_desde.strftime("%d-%m-%Y")
            else:
                fecha_desde_str = str(fecha_desde)
            if isinstance(fecha_hasta, datetime.datetime):
                fecha_hasta_str = fecha_hasta.strftime("%d-%m-%Y")
            else:
                fecha_hasta_str = str(fecha_hasta)
            valor_celda = f"Desde: {fecha_desde_str} \n Hasta: {fecha_hasta_str}"
            ws_hist.cell(row=fila_hist, column=14, value=valor_celda)
            if not fecha_desde or not fecha_hasta:
                raise ValueError("Error: No se han especificado ambas fechas para la descarga de datos.")
            break
        else:
            valor_celda = "No se descargaron de datos."

    


    print(f"✔ Datos descargados actualizados: el intervalo de fechas es {valor_celda}.")

    wb_hist.save(excel_historico)
    wb_hist.close()


def actualizar_baterias_destino(n_equipo, codigo_incidencias, fila_hist, carpeta_destino_historico, ubicacion):
    """
    Actualiza el estado de las baterías en el archivo histórico del equipo (columna 11).
    - Clasifica las baterías según su estado (SOH).
    - Si hay baterías en mal estado, lo registra.
    - Si todas están bien, también lo indica.
    """
    soh_baterias=None


    for incidencia in codigo_incidencias:
        if incidencia[0] in ("Estado_Baterias", "Error_Baterias"):
            soh_baterias = incidencia[1]
            # --- 1. Clasificar baterías ---
            baterias_buen_estado = sum(1 for soh in soh_baterias if 80 <= soh <= 100)
            baterias_mal_estado = sum(1 for soh in soh_baterias if 0 <= soh < 80)
            
            if any(soh < 0 or soh > 100 for soh in soh_baterias):
                raise ValueError(f"Uno de los estados de batería está fuera de rango (0-100).")

            total_baterias = baterias_buen_estado + baterias_mal_estado
            
            # --- 2. Preparar el mensaje de estado ---
            nuevo_estado = ""
            if baterias_mal_estado > 0:
                if "Error_Baterias" in codigo_incidencias:
                    nuevo_estado = f"{baterias_mal_estado}/{total_baterias}"
                    print("Advertencia: Baterías no están en buen estado. Se deben cambiar.")
                else:
                    # Este caso implica que se marcaron baterías para cambiar, pero no se hizo. Lo registramos.
                    nuevo_estado = f"{baterias_mal_estado}/{total_baterias}"
                    print("Advertencia: Se detectaron baterías en mal estado.")
            elif baterias_buen_estado == 0:
                nuevo_estado = f"{baterias_mal_estado}/{total_baterias}"
                print("Todas las baterías están en buen estado.")
            else:
                nuevo_estado = "No se registró el soh de las baterías."
                print("No se detectaron baterías.")

    if not soh_baterias:
        print("ℹ No se ha registrado el estado de las baterías.")
        nuevo_estado = ""

    

    # --- 3. Actualizar Excel Histórico ---
    excel_historico = os.path.join(carpeta_destino_historico, f"{n_equipo}.xlsx")
    wb_hist = load_workbook(excel_historico)
    ws_hist = wb_hist[ubicacion]

    ws_hist.cell(row=fila_hist, column=11, value=nuevo_estado)

    wb_hist.save(excel_historico)
    wb_hist.close()
    print(f"✔ Estado de baterías actualizado en el histórico para el equipo {n_equipo}.")


def actualizar_sensores_destino(informe_destino, n_equipo, codigo_incidencias, fila_hist, carpeta_destino_historico, ubicacion, hoja_destino_name='Lidar Windcube'):
    """
    Actualiza la información de sensores cambiados en:
    1. El Excel padre (hoja 'Lidar Windcube').
    2. El registro del archivo histórico (columna 12).
    
    Parámetros:
        informe_destino (str): Ruta del Excel padre.
        n_equipo (str|int): Identificador del LIDAR.
        codigo_incidencias (list): Lista de tuplas con el tipo de incidencia y sus datos.
        fila_hist (int): Número de fila en el histórico donde se está registrando este mantenimiento.
        carpeta_destino_historico (str): Carpeta donde están los excels de histórico.
        ubicacion (str): Nombre de la hoja en el histórico (ubicación del LIDAR).
        hoja_destino_name (str): Nombre de la hoja del Excel padre (default: 'Lidar Windcube').
    """
    sensores_cambiados_ultimomantenimiento = None
    wb_padre = None
    wb_hist = None

    try:
        # --- 1. Extraer datos de sensores de la lista de incidencias ---
        for incidencia in codigo_incidencias:
            if incidencia[0] == "Error_Sensores":
                sensores_cambiados_ultimomantenimiento = incidencia[1]
                break

        if not sensores_cambiados_ultimomantenimiento:
            print("ℹ No se han registrado sensores cambiados en este mantenimiento.")
            return

        # Función auxiliar para encontrar la fila del equipo, para que la función sea autónoma.
        def encontrar_fila_destino(wb, n_equipo, hoja_name):
            ws = wb[hoja_name]
            for i, row in enumerate(ws.iter_rows(min_row=3, values_only=True), 3):
                if row[0] == n_equipo:
                    return i
            return None

        # --- 2. Formatear la información de los sensores ---
        sensores_str_list = [f"{sensor} ({ns_antiguo} -> {ns_nuevo})"
                             for sensor, ns_antiguo, ns_nuevo in sensores_cambiados_ultimomantenimiento]
        sensores_cambiados_str = ", ".join(sensores_str_list)

        # --- 3. Actualizar Excel Padre ---
        wb_padre = load_workbook(informe_destino)
        hoja_padre = wb_padre[hoja_destino_name]
        
        fila_destino = encontrar_fila_destino(wb_padre, n_equipo, hoja_destino_name)
        if not fila_destino:
            print(f"Advertencia: No se encontró el equipo {n_equipo} en el archivo padre.")
            return

        comentario_ult = hoja_padre.cell(row=fila_destino, column=17).value or ""
        nuevo_comentario_padre = f"{comentario_ult} | Sensores cambiados: {sensores_cambiados_str}" if comentario_ult else f"Sensores cambiados: {sensores_cambiados_str}"
        hoja_padre.cell(row=fila_destino, column=17, value=nuevo_comentario_padre)

        wb_padre.save(informe_destino)
        print(f"✔ Información de sensores actualizada en el archivo padre para el equipo {n_equipo}.")

        # --- 4. Actualizar Excel Histórico ---
        excel_historico = os.path.join(carpeta_destino_historico, f"{n_equipo}.xlsx")
        wb_hist = load_workbook(excel_historico)
        ws_hist = wb_hist[ubicacion]

        sensores_previos = ws_hist.cell(row=fila_hist, column=12).value or ""
        valor_final = f"{sensores_previos}\n{sensores_cambiados_str}" if sensores_previos else sensores_cambiados_str
        ws_hist.cell(row=fila_hist, column=12, value=valor_final)

        wb_hist.save(excel_historico)
        
        print(f"✔ Información de sensores actualizada en el histórico para el equipo {n_equipo}.")

    except Exception as e:
        print(f"Error al actualizar la información de sensores: {e}")
        sys.exit(1)
    finally:
        if wb_padre:
            wb_padre.close()
        if wb_hist:
            wb_hist.close()



def actualizar_incidencias_destino(informe_destino, n_equipo, codigo_incidencias, fila_hist, 
                                   carpeta_destino_historico, ubicacion, 
                                   hoja_destino_name="Lidar Windcube"):
    """
    Actualiza el comentario del último mantenimiento con las incidencias de tipo "Error" en:
    1. La hoja destino del Excel padre (columna 17).
    2. El archivo histórico del equipo (columna 3).
    
    Parámetros:
        informe_destino (str): Ruta del Excel padre.
        n_equipo (str|int): Identificador del LIDAR.
        codigo_incidencias (list): Lista de tuplas con el tipo de incidencia y sus datos.
        fila_hist (int): Número de fila en el histórico donde se está registrando este mantenimiento.
        carpeta_destino_historico (str): Carpeta donde están los excels de histórico.
        ubicacion (str): Nombre de la hoja en el histórico (ubicación del LIDAR).
        hoja_destino_name (str): Nombre de la hoja del Excel padre (default: 'Lidar Windcube').
    """
    incidencias_con_error = []
    wb_padre = None
    wb_hist = None
    
    try:
        # --- 1. Extraer solo las incidencias que comienzan con "Error" ---
        for incidencia in codigo_incidencias:
            if isinstance(incidencia[0], str) and incidencia[0].startswith("Error"):
                incidencias_con_error.append(incidencia[0])

        if not incidencias_con_error:
            print("ℹ No se han registrado incidencias en este mantenimiento.")
            return

        incidencias_str = ", ".join(incidencias_con_error)

        # --- 2. Actualizar Excel Padre ---
        wb_padre = load_workbook(informe_destino)
        hoja_padre = wb_padre[hoja_destino_name]
        
        # Llamamos a la función externa para encontrar la fila del equipo
        fila_destino = encontrar_fila_destino(wb_padre, n_equipo, hoja_destino_name)
        if not fila_destino:
            print(f"Advertencia: No se encontró el equipo {n_equipo} en el archivo padre.")
            return
            
        comentario_ult = hoja_padre.cell(row=fila_destino, column=17).value or ""
        nuevo_comentario_padre = f"{comentario_ult} | Incidencias: {incidencias_str}" if comentario_ult else f"Incidencias: {incidencias_str}"
        hoja_padre.cell(row=fila_destino, column=17, value=nuevo_comentario_padre)

        wb_padre.save(informe_destino)
        print(f"✔ Incidencias añadidas correctamente al archivo padre para el equipo {n_equipo}.")

        # --- 3. Actualizar Excel Histórico ---
        excel_historico = os.path.join(carpeta_destino_historico, f"{n_equipo}.xlsx")
        wb_hist = load_workbook(excel_historico)
        ws_hist = wb_hist[ubicacion]

        comentario_hist = ws_hist.cell(row=fila_hist, column=3).value or ""
        nuevo_comentario_hist = f"{comentario_hist} | Incidencias: {incidencias_str}" if comentario_hist else f"Incidencias: {incidencias_str}"
        ws_hist.cell(row=fila_hist, column=3, value=nuevo_comentario_hist)

        wb_hist.save(excel_historico)
        print(f"✔ Incidencias añadidas correctamente al histórico para el equipo {n_equipo}.")
        
    except Exception as e:
        print(f"Error al actualizar la información de incidencias: {e}")
        sys.exit(1)
    finally:
        if wb_padre:
            wb_padre.close()
        if wb_hist:
            wb_hist.close()


# === Funcion de envío de email ===

def enviar_correo(destinatario, asunto, mensaje, hora_envio=None, cc=None, adjuntos=None):
    """
    Programar el envío de un correo usando Outlook instalado en Windows.
    
    Parámetros:
    - destinatario: str o lista de correos
    - asunto: str
    - mensaje: str (texto plano)
    - hora_envio: datetime.datetime o pandas.Timestamp (opcional, hora programada)
    - cc: str o lista de correos (opcional)
    - adjuntos: lista de rutas de archivos (opcional)
    """
    outlook = win32.Dispatch('Outlook.Application')
    mail = outlook.CreateItem(0)  # 0 = MailItem

    # Destinatarios
    mail.To = ";".join(destinatario) if isinstance(destinatario, list) else destinatario

    # CC
    if cc:
        mail.CC = ";".join(cc) if isinstance(cc, list) else cc

    # Asunto y cuerpo
    mail.Subject = asunto
    mail.Body = mensaje

    # Adjuntos
    if adjuntos:
        for archivo in adjuntos:
            if os.path.isfile(archivo):
                mail.Attachments.Add(archivo)

    # Guardar los destinatarios antes de enviar
    destinatarios_str = mail.To

    # Programar envío
    if hora_envio:
        if isinstance(hora_envio, pd.Timestamp):
            hora_envio = hora_envio.to_pydatetime()

        # Si no se especifica hora, añadirle la zona local de Windows 
        if hora_envio.tzinfo is None:
            local_tz = get_localzone()  # devuelve un ZoneInfo
            hora_envio = hora_envio.replace(tzinfo=local_tz)

        mail.DeferredDeliveryTime = hora_envio

    mail.Send()
    print(f"✅ Correo se enviará a {destinatarios_str}" + (f" a las {hora_envio}" if hora_envio else ""))
