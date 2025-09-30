# mantenimiento.py

import pandas as pd
import numpy as np
import sys
from openpyxl import load_workbook
from datetime import datetime

# Leer archivo de origen con la siguiente función:

def leer_datos_origen(informe_mtto, hoja_mtto):
    """
    Lee los datos del archivo de origen, y devuelve un DataFrame si y solo si los datos tienen el formsto correcto.
    """
    df = pd.read_excel(informe_mtto, sheet_name=hoja_mtto, engine='openpyxl')


    # Comprobar que el DataFrame tiene al menos una fila y las columnas necesarias
    if df.empty or df.shape[1] < 19:
        raise ValueError("El DataFrame está vacío o no tiene suficientes columnas.")
    

    n_equipo = df.iloc[0, 0]  # Obtiene valor n equipo LiDAR
    fecha_ultimomantenimiento = df.iloc[0, 1]  # Obtiene valor fecha del ultimo mantenimiento
    comentario_ultimomantenimiento = df.iloc[0, 16]  # Obtiene valor comentario del ultimo mantenimiento


     # Obtenemos booleano: Se añade liquido limpiaparabrisas?
    adicion_liquido = df.iloc[0, 5]
    assert isinstance(adicion_liquido, (int, np.integer, float, np.floating)), \
        f"Liquido añadido debe ser int o float, no {type(adicion_liquido)}"
    adicion_liquido = False if pd.isna(adicion_liquido) or adicion_liquido == 0 else True
    # Obtiene valor liquido parabrisas del ultimo mantenimiento [Cambio liquido (booleano), Liquido añadido (int), Liquido restante (int)]  
    liquido_ultimomantenimiento = [adicion_liquido , df.iloc[0, 5], df.iloc[0, 6]] 

    # Obtenemos booleano: Se añade metanol?
    adicion_metanol = df.iloc[0, 7]
    assert isinstance(adicion_metanol, (int, np.integer, float, np.floating)), \
        f"Metanol añadido debe ser numérico, no {type(adicion_metanol)}"
    adicion_metanol = False if pd.isna(adicion_metanol) or adicion_metanol == 0 else True
    # Obtiene valor metanol del ultimo mantenimiento [Cambio metanol (booleano), Cartuchos añadidos al EFOY (int), Cartuchos añadidos al STOCK (int)]
    metanol_ultimomantenimiento = [adicion_metanol , df.iloc[0, 7], df.iloc[0, 8]]
    


    # Normalizamos los valores booleanos de las columnas que contienen 'SI'/'NO'
    columnas_si_no = [2, 3, 4, 9, 10, 11, 12] # Columnas que contienen 'SI'/'NO'
    for col in columnas_si_no:
        valor = str(df.iloc[0, col]).strip().upper()  # normaliza texto
        if valor == "SI":
            df.iloc[0, col] = True
        elif valor == "NO":
            df.iloc[0, col] = False
        else:
            raise ValueError(f"Valor inesperado en la columna {col}: {valor}. Debe ser 'SI' o 'NO'.")
    
    # Comprobar si hay incidencias y generar un parte de incidencias:
    # Si alguna de las columnas de estado general del LiDAR es False, se genera un parte de incidencias.
    codigo_incidencias = []

    if df.iloc[0, 2]==False: # EFOY funciona?
        print("Advertencia: EFOY no funciona")
        codigo_incidencias.append("Error_EFOY"),
    
    if df.iloc[0, 3]==False: # Filtros de Lidar correctos?
        print("Advertencia: Filtros de Lidar no estan correctos")
        filtro_desechado = str(df.iloc[2, 3]).strip().upper()  # Filtro desechado?
        if filtro_desechado=="SI": 
            codigo_incidencias.append("Error_Filtro_Desechado")
        elif filtro_desechado=="NO":
            codigo_incidencias.append("Error_Filtro_Cambiado")
        else:
            print("Advertencia: No se ha especificado si el filtro de Lidar esta desechado o no")
            sys.exit(1)

    if df.iloc[0, 4]==False: # Escobilla limpia?
        print("Advertencia: Escobilla no esta limpia")
        codigo_incidencias.append("Error_Escobilla")

    if df.iloc[0, 9]==True: # Se cambian las baterías? Creamos una lista de estados de baterías:
        print("Advertencia: Se han cambiado las baterías")
        codigo_incidencias.append("Error_Baterias")
        estado_baterias = df.iloc[1:,9].dropna().tolist()  # Lista de estados de baterías
        # Comprobar que estado_baterias es una lista de porcentajes (int o float):
        if not all(isinstance(x, (int, float)) for x in estado_baterias):
            raise ValueError("Los estados de baterías deben ser números (int o float).")
        if not estado_baterias:
            print("Advertencia: No se ha registrado el estado de las baterías")
            estado_baterias = []
    else:
        estado_baterias = df.iloc[1:,9].dropna().tolist()  # Lista de estados de baterías
        # Comprobar que estado_baterias es una lista de porcentajes (int o float):
        if not all(isinstance(x, (int, float)) for x in estado_baterias):
            raise ValueError("Los estados de baterías deben ser números (int o float).")
        if not estado_baterias:
            print("Advertencia: No se han registrado estados de baterías")
            estado_baterias = []
        print("No se cambian las baterías, pero se han registrado sus estados.")

    if df.iloc[0, 10]==False: # Bomba de agua funciona?
        print("Advertencia: Bomba de agua no funciona")
        codigo_incidencias.append("Error_Bomba")

    if df.iloc[0, 11]==False: # Extintor de incendios revisado?
        print("Advertencia: Extintor de incendios no esta revisado")
        codigo_incidencias.append("Error_Extintor")

    if df.iloc[0, 12]==False: # Descarga de datos?
        print("Advertencia: No se han descargado los datos")
        codigo_incidencias.append("Error_DescargaDatos")
    
    if df.iloc[0, 13]: # Sensores cambiados?
        # Contar cuantas columnas de sensores cambiados hay:
        sensores_cambiados_ultimomantenimiento = df.iloc[0:, 13].dropna().tolist()
        if len(sensores_cambiados_ultimomantenimiento) > 0:
            print("Se han cambiado los siguientes sensores:", sensores_cambiados_ultimomantenimiento)
            nserie_cambio = df.iloc[0:, 14].dropna().tolist()
            nserie_recambio = df.iloc[0:, 15].dropna().tolist()
            sensores_cambiados_ultimomantenimiento = list(zip(sensores_cambiados_ultimomantenimiento, nserie_cambio, nserie_recambio))
            codigo_incidencias.append("Error_Sensores")
        else:
            sensores_cambiados_ultimomantenimiento = []

        # Si se cambia algún sensor, obtenemos el nombre del sensor cambiado y el de recambio:
    
    # Si no hay incidencias, se añade un mensaje de que todo esta correcto:
    if not codigo_incidencias:
        print("No hay incidencias, todo correcto.")

    # Si se cambia algún sensor, obtenemos el nombre del sensor cambiado y el de recambio:
   

    # Comprueba el formato de los datos leídos:
    assert isinstance(n_equipo, (str, int)), f"n_equipo debe ser str o int, no {type(n_equipo)}"
    assert pd.api.types.is_datetime64_any_dtype(df.iloc[:, 1]), f"fecha_ultimomantenimiento debe ser datetime, no {type(fecha_ultimomantenimiento)}"
    assert isinstance(comentario_ultimomantenimiento, str), f"comentario_ultimomantenimiento debe ser str, no {type(comentario_ultimomantenimiento)}"
    assert isinstance(liquido_ultimomantenimiento[0], (bool, int)), f"Cambio liquido debe ser bool o int, no {type(liquido_ultimomantenimiento[0])}"
    assert isinstance(liquido_ultimomantenimiento[1],(int, np.integer, float, np.floating)), f"Liquido añadido debe ser int o float, no {type(liquido_ultimomantenimiento[1])}"
    assert isinstance(liquido_ultimomantenimiento[2],(int, np.integer, float, np.floating)), f"Liquido restante debe ser int o float, no {type(liquido_ultimomantenimiento[2])}"
    assert isinstance(metanol_ultimomantenimiento[0], (bool, int)), f"Cambio metanol debe ser bool o int, no {type(metanol_ultimomantenimiento[0])}"
    assert isinstance(metanol_ultimomantenimiento[1],(int, np.integer, float, np.floating)), f"Cartuchos añadidos al EFOY debe ser int o float, no {type(metanol_ultimomantenimiento[1])}"
    assert isinstance(metanol_ultimomantenimiento[2],(int, np.integer, float, np.floating)), f"Cartuchos añadidos al STOCK debe ser int o float, no {type(metanol_ultimomantenimiento[2])}"
    assert isinstance(df.iloc[0, 2], bool), f"EFOY funciona debe ser SI o NO, no {type(df.iloc[0, 2])}"
    assert isinstance(df.iloc[0, 3], bool), f"Filtros de Lidar correctos debe ser SI o NO, no {type(df.iloc[0, 3])}"
    assert isinstance(df.iloc[0, 4], bool), f"Escobilla limpia debe ser SI o NO, no {type(df.iloc[0, 4])}"
    assert isinstance(df.iloc[0, 9], bool), f"Se cambian las baterías? debe ser SI o NO, no {type(df.iloc[0, 9])}"
    assert isinstance(df.iloc[0, 10], bool), f"Bomba de agua funciona debe ser SI o NO, no {type(df.iloc[0, 10])}"
    assert isinstance(df.iloc[0, 11], bool), f"Extintor de incendios revisado debe ser SI o NO, no {type(df.iloc[0, 11])}"
    assert isinstance(df.iloc[0, 12], bool), f"Descarga de datos debe ser SI o NO, no {type(df.iloc[0, 12])}"

    assert isinstance(sensores_cambiados_ultimomantenimiento, list), f"sensores_cambiados_ultimomantenimiento debe ser una lista, no {type(sensores_cambiados_ultimomantenimiento)}"



    print("Datos leídos correctamente del archivo de origen, para el número de equipo:", n_equipo)

    return n_equipo, fecha_ultimomantenimiento, comentario_ultimomantenimiento, metanol_ultimomantenimiento, liquido_ultimomantenimiento, codigo_incidencias , sensores_cambiados_ultimomantenimiento, estado_baterias






#Actualizar el archivo destino con los datos leídos del archivo de origen, a partir de las siguientes funciones:

def actualizar_fecha_destino(informe_destino, n_equipo, fecha_ultimomantenimiento, hoja_destino_name='Lidar Windcube', hoja_destino_historico_name='Historico'):
    """
    Actualiza el archivo destino con la fecha del archivo de origen.
    """
    # Cargar el libro de trabajo
    wb = load_workbook(informe_destino)
    hoja_destino = wb[hoja_destino_name]
    hoja_destino_historico = wb[hoja_destino_historico_name]

    # Buscar la fila donde la primera columna coincide con n_equipo LiDAR,
    # copiar la fecha de ultimo mantenimiento desde el archivo de origen y hacer append en la columna 2 de la hoja 'Historico',
    # y actualizar la columna 5 con la nueva fecha de ultimo mantenimiento:
    for idx, row in enumerate(hoja_destino.iter_rows(min_row=3, max_col=1, values_only=True), start=3):
        if row[0] == n_equipo:
            # Accede fecha del penultimo mantenimiento, presente en el excel padre
            fecha_penultimomantenimiento = hoja_destino.cell(row=idx, column=5).value

            # Formatear fechas para que siempre sean dd/mm/yyyy
            fecha_penult_str = fecha_penultimomantenimiento.strftime("%d/%m/%Y") if isinstance(fecha_penultimomantenimiento, datetime) else str(fecha_penultimomantenimiento)
            fecha_ult_str   = fecha_ultimomantenimiento.strftime("%d/%m/%Y") if isinstance(fecha_ultimomantenimiento, datetime) else str(fecha_ultimomantenimiento)
            if fecha_penult_str == fecha_ult_str:
                print(f"Advertencia: La fecha del último mantenimiento ({fecha_ult_str}) es la misma que la del penúltimo mantenimiento. No se actualizará.")
                sys.exit(1)
            # Actualizar la columna 5 con la fecha del último mantenimiento
            hoja_destino.cell(row=idx, column=5, value=fecha_ultimomantenimiento)
            # Concatenar la fecha del último mantenimiento en la hoja 'Historico'
            historico_fechas = hoja_destino_historico.cell(row=idx, column=2).value 

            # Esta función toma un valor que puede ser una fecha, una cadena de fechas separadas   
            # por comas, o None, y devuelve una cadena con las fechas formateadas.
            def formatear_varias_fechas(valor):
                if valor is None:
                    return ""
                if isinstance(valor, datetime):
                    return valor.strftime("%d/%m/%Y")
                if isinstance(valor, str):
                    partes = [p.strip() for p in valor.split(",")]
                    nuevas_partes = []
                    for parte in partes:
                        try:
                            fecha = datetime.strptime(parte, "%Y-%m-%d %H:%M:%S")
                            nuevas_partes.append(fecha.strftime("%d/%m/%Y"))
                        except ValueError:
                            nuevas_partes.append(parte)  # No era una fecha con ese formato
                    return ", \n ".join(nuevas_partes) # Une elementos de una lista de strings en una sola string separada por comas y enters
                return str(valor)
            
            # Formatear las fechas del historico
            historico_fechas = formatear_varias_fechas(historico_fechas)

            if historico_fechas:
                # Si ya hay un valor, concatenar como una cadena separada por comas
                nuevo_historico_fechas = f"{historico_fechas}, \n {fecha_ult_str}"
            else:
                # Sino, iniciar con la fecha del último mantenimiento
                nuevo_historico_fechas = str(fecha_ult_str)

            hoja_destino_historico.cell(row=idx, column=2, value=nuevo_historico_fechas)
            break
    # Guardar los cambios en el archivo destino
    wb.save(informe_destino)
    wb.close()




def actualizar_comentario_destino(informe_destino, n_equipo, comentario_ultimomantenimiento, hoja_destino_name='Lidar Windcube', hoja_destino_historico_name='Historico'):
    """
    Actualiza el archivo destino con el comentario del archivo de origen.
    """
    # Cargar el libro de trabajo
    wb = load_workbook(informe_destino)
    hoja_destino = wb[hoja_destino_name]
    hoja_destino_historico = wb[hoja_destino_historico_name]

    # Buscar la fila donde la primera columna coincide con n_equipo LiDAR,
    # copiar el comentario del ultimo mantenimiento desde el archivo de origen, hacer append en la columna 3 de la hoja 'Historico',
    # y actualizar la columna 8 con el comentario del ultimo mantenimiento:
    for idx, row in enumerate(hoja_destino.iter_rows(min_row=3, max_col=1, values_only=True), start=3):
        if row[0] == n_equipo:
            # Actualizar la columna 8 con el comentario del ultimo mantenimiento
            hoja_destino.cell(row=idx, column=8, value=comentario_ultimomantenimiento)
            # Concatenar el comentario del último mantenimiento en la hoja 'Historico'
            historico_comentarios = hoja_destino_historico.cell(row=idx, column=3).value 
            if historico_comentarios:
                # Si ya hay un valor, concatenar como una cadena separada por comas
                nuevo_historico_comentarios = f"{historico_comentarios}, \n {comentario_ultimomantenimiento}"
            else:
                # Sino, iniciar con el comentario del último mantenimiento
                nuevo_historico_comentarios = str(comentario_ultimomantenimiento)

            hoja_destino_historico.cell(row=idx, column=3, value=nuevo_historico_comentarios)
            break
    # Guardar los cambios en el archivo destino
    wb.save(informe_destino)
    wb.close()




def actualizar_metanol_destino(informe_destino, n_equipo, metanol_ultimomantenimiento, hoja_destino_name='Lidar Windcube', hoja_destino_historico_name='Historico'):
    """
    Actualiza las columnas 4 y 5 del archivo destino con el metanol del archivo de origen.
    """
    # Cargar el libro de trabajo
    wb = load_workbook(informe_destino)
    hoja_destino = wb[hoja_destino_name]
    hoja_destino_historico = wb[hoja_destino_historico_name]
    cambio_metanol = metanol_ultimomantenimiento[0]  # Booleano que indica si se ha cambiado el metanol
    metanol_anadido_EFOY = metanol_ultimomantenimiento[1]  # Cantidad de metanol añadido al EFOY
    metanol_anadido_stock = metanol_ultimomantenimiento[2]  # Cantidad de metanol añadido al stock



    # Buscar la fila donde la primera columna coincide con n_equipo LiDAR,
    # Y actualizar el metanol si y solo si se ha cambiado:
    if cambio_metanol:  # Si se ha cambiado el metanol
        for idx, row in enumerate(hoja_destino.iter_rows(min_row=3, max_col=1, values_only=True), start=3):
            if row[0] == n_equipo:
                # Actualizar la columna 4 con el valor de cartuchos de metanol en STOCK tras el mantenimiento
                metanol_stock_restante = hoja_destino_historico.cell(row=idx, column=4).value # Metanol en stock antes del mantenimiento
                hoja_destino_historico.cell(row=idx, column=4, value= metanol_stock_restante + metanol_anadido_stock - metanol_anadido_EFOY )  # Actualizamos el metanol en stock despues el mantenimiento
                # Actualizar la columna 5 con el valor de cartuchos de metanol gastados en la historia del LiDAR (= Cartuchos totales instalados en el EFOY)
                metanol_usado = hoja_destino_historico.cell(row=idx, column=5).value # Metanol usado por el LiDAR hasta el mantenimiento anterior
                hoja_destino_historico.cell(row=idx, column=5, value= metanol_usado + metanol_anadido_EFOY )  # Actualizamos el metanol usado por el EFOY, despues el mantenimiento
                print("En este mantenimiento se cambio el metanol")
                break

    elif cambio_metanol == False:
        print("En este mantenimiento no se cambio el metanol")
                

    # Guardar los cambios en el archivo destino
    wb.save(informe_destino)
    wb.close()



def actualizar_liquido_destino(informe_destino, n_equipo, liquido_ultimomantenimiento, hoja_destino_name='Lidar Windcube', hoja_destino_historico_name='Historico'):
    """
    Actualiza las columnas 6 y 7 del archivo destino con el liquido del archivo de origen.
    """
    # Cargar el libro de trabajo
    wb = load_workbook(informe_destino)
    hoja_destino = wb[hoja_destino_name]
    hoja_destino_historico = wb[hoja_destino_historico_name]
    adicion_liquido = liquido_ultimomantenimiento[0]  # Booleano que indica si se ha cambiado el liquido
    liquido_anadido = liquido_ultimomantenimiento[1]  # Cantidad de liquido añadido al deposito de liquido limpiaparabrisas
    liquido_restante = liquido_ultimomantenimiento[2]  # Cantidad de liquido restante en el deposito de liquido limpiaparabrisas al llegar al mantenimiento



    # Buscar la fila donde la primera columna coincide con n_equipo LiDAR,
    # Y actualizar el liquido si y solo si se ha cambiado:
    if adicion_liquido:  # Si se ha cambiado el liquido
        for idx, row in enumerate(hoja_destino.iter_rows(min_row=3, max_col=1, values_only=True), start=3):
            if row[0] == n_equipo:
                # Actualizar la columna 6 con el valor de liquido restante tras el mantenimiento
                # liquido_restante_penultimomantenimineto = hoja_destino_historico.cell(row=idx, column=6).value # Liquido restante antes al llegar en el penulmantenimiento
                hoja_destino_historico.cell(row=idx, column=6, value= liquido_restante + liquido_anadido)  # Actualizamos el liquido restante despues el mantenimiento
                # Actualizar la columna 7 con el valor de litros de liquido gastados en la historia del LiDAR (= Litros totales echados en el deposito de liquido limpiaparabrisas)
                liquido_usado = hoja_destino_historico.cell(row=idx, column=7).value # Liquido usado por el LiDAR hasta el mantenimiento anterior
                hoja_destino_historico.cell(row=idx, column=7, value = liquido_usado + liquido_anadido)  # Actualizamos el liquido usado por el limpiaparabrisas, despues el mantenimiento
                # !!!!!!!!!!!!!!!!!!!
                # Podría cambiar el liquido usado a: value = liquido_usado + (liquido_restante_penultimomantenimineto - (liquido_restante+liquido_anadido))
                print("En este mantenimiento se añadio liquido limpiaparabrisas")
                break

    elif adicion_liquido == False:
        for idx, row in enumerate(hoja_destino.iter_rows(min_row=3, max_col=1, values_only=True), start=3):
            if row[0] == n_equipo:
                # Actualizar la columna 6 con el valor de liquido restante tras el mantenimiento
                # liquido_restante_penultimomantenimineto = hoja_destino_historico.cell(row=idx, column=6).value # Liquido restante antes del mantenimiento, al llegar
                hoja_destino_historico.cell(row=idx, column=6, value= liquido_restante)  # Actualizamos el liquido restante despues el mantenimiento
                print("En este mantenimiento no se añadio liquido limpiaparabrisas")
                break

    # Guardar los cambios en el archivo destino
    wb.save(informe_destino)
    wb.close()




def actualizar_filtros_destino(informe_destino, n_equipo, codigo_incidencias, hoja_destino_name='Lidar Windcube', hoja_destino_historico_name='Historico'):
    """
    Actualiza los filtros del archivo destino con los datos del archivo de origen.
    """
    # Cargar el libro de trabajo
    wb = load_workbook(informe_destino)
    hoja_destino = wb[hoja_destino_name]
    hoja_destino_historico = wb[hoja_destino_historico_name]

    # Buscar la fila donde la primera columna coincide con n_equipo LiDAR,
    # y actualizar los contadore de filtros cambiados y desechados según el código de incidencias:
    for idx, row in enumerate(hoja_destino.iter_rows(min_row=3, max_col=1, values_only=True), start=3):
        if row[0] == n_equipo:
            # Actualizar la columna 8 con el contador de filtros cambiados
            contador_filtros_cambiados = hoja_destino_historico.cell(row=idx, column=8).value or 0
            if "Error_Filtro_Cambiado" in codigo_incidencias:
                hoja_destino_historico.cell(row=idx, column=8, value=contador_filtros_cambiados + 1)

            # Actualizar la columna 9 con el contador de filtros desechados
            contador_filtros_desechados = hoja_destino_historico.cell(row=idx, column=9).value or 0
            if "Error_Filtro_Desechado" in codigo_incidencias:
                hoja_destino_historico.cell(row=idx, column=9, value=contador_filtros_desechados + 1)

            break
            
    # Guardar los cambios en el archivo destino
    wb.save(informe_destino)
    wb.close()



def actualizar_escobilla_destino(informe_destino, n_equipo, codigo_incidencias, hoja_destino_name='Lidar Windcube', hoja_destino_historico_name='Historico'):
    """
    Actualiza los filtros del archivo destino con los datos del archivo de origen.
    """
    # Cargar el libro de trabajo
    wb = load_workbook(informe_destino)
    hoja_destino = wb[hoja_destino_name]
    hoja_destino_historico = wb[hoja_destino_historico_name]

    # Buscar la fila donde la primera columna coincide con n_equipo LiDAR,
    # y actualizar la columna 10 con el contador de escobillas cambiadas según el código de incidencias:
    for idx, row in enumerate(hoja_destino.iter_rows(min_row=3, max_col=1, values_only=True), start=3):
        if row[0] == n_equipo:
            # Actualizar la columna 10 con el contador de escobillas cambiadas
            contador_escobillas_cambiadas = hoja_destino_historico.cell(row=idx, column=10).value or 0
            if "Error_Escobilla" in codigo_incidencias:
                hoja_destino_historico.cell(row=idx, column=10, value=contador_escobillas_cambiadas + 1)

            break
    # Guardar los cambios en el archivo destino
    wb.save(informe_destino)
    wb.close()







def actualizar_incidencias_destino(informe_destino, n_equipo, codigo_incidencias, hoja_destino_name='Lidar Windcube', hoja_destino_historico_name='Historico'):

    wb = load_workbook(informe_destino)
    hoja_destino = wb[hoja_destino_name]
    hoja_hist = wb[hoja_destino_historico_name]

    # Construir el string de incidencias (si hay)
    incidencias_str = ""
    if codigo_incidencias:
        incidencias_str = "Incidencias: " + ", ".join(codigo_incidencias)

    # Buscar fila por n_equipo en la hoja destino
    fila = None
    for idx, (val,) in enumerate(hoja_destino.iter_rows(min_row=3, max_col=1, values_only=True), start=3):
        if val == n_equipo:
            fila = idx
            break
    if fila is None:
        wb.close()
        raise ValueError(f"No se encontró el equipo {n_equipo} en '{hoja_destino_name}'")

    # Actualizar comentario en hoja destino (columna 8)
    if incidencias_str:
        com = hoja_destino.cell(row=fila, column=8).value or ""
        com = (com + " | " if com else "") + incidencias_str
        hoja_destino.cell(row=fila, column=8, value=com)

    # Actualizar comentario en hoja histórico (columna 3)
    if incidencias_str:
        com_h = hoja_hist.cell(row=fila, column=3).value or ""
        com_h = (com_h + " | " if com_h else "") + incidencias_str
        hoja_hist.cell(row=fila, column=3, value=com_h)

    wb.save(informe_destino)
    wb.close()







def actualizar_baterias_destino(informe_destino, n_equipo, codigo_incidencias, estado_baterias,
                                hoja_destino_name='Lidar Windcube',
                                hoja_destino_historico_name='Historico'):
    """
    Actualiza el estado de las baterías en el archivo destino e histórico (columna 11).
    - Cuenta cuántas están >=80 (buen estado) y cuántas <80 (mal estado).
    - Si hay alguna en mal estado:
        * Si "Error_Baterias" está en codigo_incidencias -> "Se deben cambiar".
        * Si no está -> "Se han cambiado".
    - Si todas están bien -> "X de Y baterías en buen estado".
    """

    wb = load_workbook(informe_destino)
    hoja_destino = wb[hoja_destino_name]
    hoja_hist = wb[hoja_destino_historico_name]

    # Contabilizar y validar
    baterias_buen_estado = 0
    baterias_mal_estado = 0
    fuera_rango = []

    for soh in estado_baterias:
        try:
            v = float(soh)
        except (TypeError, ValueError):
            fuera_rango.append(soh)
            continue
        if 80 <= v <= 100:
            baterias_buen_estado += 1
        elif 0 <= v < 80:
            baterias_mal_estado += 1
        else:
            fuera_rango.append(soh)

    if fuera_rango:
        wb.close()
        raise ValueError(f"Valores de SOH fuera de rango o no numéricos: {fuera_rango} (esperado 0-100).")

    if baterias_buen_estado == 0 and baterias_mal_estado == 0:
        print("Advertencia: No se ha registrado el estado de las baterías")
        wb.close()
        return

    # Encontrar fila del equipo (asumiendo que ambas hojas están alineadas por fila)
    fila = None
    for idx, (val,) in enumerate(hoja_destino.iter_rows(min_row=3, max_col=1, values_only=True), start=3):
        if val == n_equipo:
            fila = idx
            break
    if fila is None:
        wb.close()
        raise ValueError(f"No se encontró el equipo {n_equipo} en '{hoja_destino_name}'")

    total = baterias_buen_estado + baterias_mal_estado

    # Mensaje resumen según el estado
    if baterias_mal_estado > 0:
        # Opcional: si hay malas y aún no existe el código, lo añadimos
        # if "Error_Baterias" not in codigo_incidencias:
        #     codigo_incidencias.append("Error_Baterias")

        accion = "Se deben cambiar" if "Error_Baterias" in codigo_incidencias else "Se han cambiado"
        resumen = f"{baterias_mal_estado} de {total} baterías en mal estado. {accion}"
    else:
        resumen = f"{baterias_buen_estado} de {total} baterías en buen estado"

    # Escribir en hoja destino (columna 11)
    hoja_destino.cell(row=fila, column=11, value=resumen)

    # Añadir/concatenar en histórico (columna 11)
    estados_baterias_pasado = hoja_hist.cell(row=fila, column=11).value
    if estados_baterias_pasado:
        nuevo_hist = f"{estados_baterias_pasado}, \n {resumen}"
    else:
        nuevo_hist = resumen
    hoja_hist.cell(row=fila, column=11, value=nuevo_hist)

    wb.save(informe_destino)
    wb.close()

