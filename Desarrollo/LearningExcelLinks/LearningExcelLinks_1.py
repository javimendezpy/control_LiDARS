import pandas as pd

from openpyxl import load_workbook

#Definimos nombre del archivo de origen y destino:
informe_mtto = r'\\S31889004.services.dekra.com\DKI\EOLICA\MAT REF\BASE DE DATOS ESTACIONES\LearningExcelLinks\informe_mtto.xlsx'
informe_destino = r'\\S31889004.services.dekra.com\DKI\EOLICA\MAT REF\BASE DE DATOS ESTACIONES\LearningExcelLinks\ExcelPadre.xlsx'

# Leer archivo de origen
df_origen = pd.read_excel(informe_mtto, sheet_name='Hoja1', engine='openpyxl')


# Obtener valor deseado
n_equipo = df_origen.iloc[0, 0] #Obtiene valor nº equipo LiDAR
fecha_ultimomantenimiento = df_origen.iloc[0, 1] #Obtiene valor fecha del ultimo mantenimiento

# Abrir el archivo destino
wb = load_workbook(informe_destino) 
hoja_destino = wb['Lidar Windcube']


# Buscar la fila donde la primera columna coincide con n_equipo LiDAR,
# y actualizar la columna 5 con la nueva fecha de ultimo mantenimiento:
for idx, row in enumerate(hoja_destino.iter_rows(min_row=3, max_col=1, values_only=True), start=3):
    if row[0] == n_equipo:
        hoja_destino.cell(row=idx, column=5, value=fecha_ultimomantenimiento)
        break


# Guardar los cambios en el archivo destino
wb.save(informe_destino)

# Cerrar el libro de trabajo
wb.close()

