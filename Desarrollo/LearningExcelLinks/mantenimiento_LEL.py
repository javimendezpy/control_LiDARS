"""
===============================================================================
 MÓDULO: mantenimiento.py
===============================================================================
Este módulo permite actualizar un archivo Excel con información de mantenimientos
de equipos LiDAR Windcube. Está diseñado para trabajar con 3 hojas principales:

    - Hoja origen: contiene los datos del último mantenimiento.
    - Hoja destino (por defecto: "Lidar Windcube")
    - Hoja histórico (por defecto: "Historico")

Cada función abre el archivo Excel, localiza la fila correspondiente al equipo
(`n_equipo`) y actualiza la información según los datos del último mantenimiento,
leyendo de la hoja origen.

-------------------------------------------------------------------------------
FUNCIONES AUXILIARES:
-------------------------------------------------------------------------------
- encontrar_fila_destino(wb, n_equipo, hoja_destino_name):
    Devuelve el número de fila en la hoja destino para un equipo concreto.
    Busca en la primera columna a partir de la fila 3.

- encontrar_fila_historico(wb, n_equipo, hoja_destino_historico_name):
    Devuelve el número de fila en la hoja histórico para un equipo concreto.
    Busca en la primera columna a partir de la fila 3.

-------------------------------------------------------------------------------
FUNCIONES PRINCIPALES:
-------------------------------------------------------------------------------
- leer_datos_origen(informe_mtto, hoja_mtto):
    Lee los datos del archivo de origen y devuelve un DataFrame con los datos
    del último mantenimiento. Comprueba que el formato es correcto y que las
    columnas necesarias están presentes.

- actualizar_fecha_destino(informe_destino, n_equipo, fecha_ultimomantenimiento):
    Actualiza la fecha del último mantenimiento en la hoja destino y añade
    la nueva fecha al histórico (columna 2). Si la fecha es igual a la anterior,
    emite advertencia y detiene el programa.

- actualizar_comentario_destino(informe_destino, n_equipo, comentario_ultimomantenimiento):
    Sustituye el comentario en la hoja destino (columna 8) y mueve el comentario
    previo al histórico (columna 3).

- actualizar_metanol_destino(informe_destino, n_equipo, metanol_ultimomantenimiento):
    Actualiza en el histórico (columnas 4 y 5) el stock de metanol restante y la
    cantidad total de metanol usada. Los datos se reciben como una tupla:
    (cambio_metanol, metanol_anadido_EFOY, metanol_anadido_stock).

- actualizar_liquido_destino(informe_destino, n_equipo, liquido_ultimomantenimiento):
    Actualiza en el histórico (columnas 6 y 7) el stock de líquido limpiaparabrisas
    y la cantidad total usada. Los datos se reciben como una tupla:
    (adicion_liquido, liquido_anadido, liquido_restante).

- actualizar_filtros_destino(informe_destino, n_equipo, codigo_incidencias):
    Actualiza el estado de los filtros en el histórico (columnas 8 y 9) según
    las incidencias detectadas. Incrementa contadores de filtros cambiados y
    desechados.

- actualizar_escobilla_destino(informe_destino, n_equipo, codigo_incidencias):
    Actualiza el estado de la escobilla en el histórico (columna 10) según
    las incidencias detectadas. Incrementa el contador de escobillas cambiadas.

- actualizar_incidencias_destino(informe_destino, n_equipo, codigo_incidencias):
    Añade las incidencias detectadas al comentario del último mantenimiento
    (columna 8 de la hoja destino) y también al histórico (columna 3).

- actualizar_baterias_destino(informe_destino, n_equipo, codigo_incidencias, estado_baterias):
    Actualiza el estado de las baterías en el histórico (columna 11) según
    las incidencias detectadas. Si se cambiaron baterías, incrementa el contador
    de baterías cambiadas.

- actualizar_sensores_destino(informe_destino, n_equipo, sensores_cambiados_ultimomantenimiento):
    Actualiza el estado de los sensores en el histórico (columna 12) según
    los sensores cambiados en el último mantenimiento. 

    
-------------------------------------------------------------------------------
NOTAS IMPORTANTES:
-------------------------------------------------------------------------------
1. Todas las funciones guardan y cierran el archivo Excel tras realizar cambios.
2. Los nombres de hojas pueden personalizarse pasando los argumentos
   `hoja_destino_name` y `hoja_destino_historico_name`.
3. En caso de errores graves (por ejemplo, fechas duplicadas o valores fuera
   de rango), se usa `sys.exit(1)` o se lanza `ValueError`.
4. Las funciones están diseñadas para ser utilizadas en un flujo de trabajo
    donde se leen datos de un archivo de mantenimiento y se actualizan en un
    archivo maestro de seguimiento de equipos.

===============================================================================
"""

# Importar las librerías necesarias
from openpyxl import load_workbook # Para trabajar con archivos Excel
from datetime import datetime # Para manejar fechas
import sys # Para manejar errores y salir del programa
import pandas as pd # Para manejar DataFrames
import numpy as np # Para manejar arrays y operaciones numéricas



# === Funciones principales: Lectura de datos ===
def leer_datos_origen(informe_mtto, hoja_mtto):
    """
    Lee los datos del archivo de origen, y devuelve un DataFrame si y solo si los datos tienen el formsto correcto.
    """
    df = pd.read_excel(informe_mtto, sheet_name=hoja_mtto, engine='openpyxl')


    # Comprobar que el DataFrame tiene al menos una fila y las columnas necesarias
    if df.empty or df.shape[1] < 19:
        raise ValueError("El DataFrame está vacío o no tiene suficientes columnas.")
    

    n_equipo = df.iloc[0, 0]  # Obtiene valor n equipo LiDAR
    fecha_ultimomantenimiento = df.iloc[0, 1]  # Obtiene valor fecha del ultimo mantenimiento
    comentario_ultimomantenimiento = df.iloc[0, 16]  # Obtiene valor comentario del ultimo mantenimiento


     # Obtenemos booleano: Se añade liquido limpiaparabrisas?
    adicion_liquido = df.iloc[0, 5]
    assert isinstance(adicion_liquido, (int, np.integer, float, np.floating)), \
        f"Liquido añadido debe ser int o float, no {type(adicion_liquido)}"
    adicion_liquido = False if pd.isna(adicion_liquido) or adicion_liquido == 0 else True
    # Obtiene valor liquido parabrisas del ultimo mantenimiento [Cambio liquido (booleano), Liquido añadido (int), Liquido restante (int)]  
    liquido_ultimomantenimiento = [adicion_liquido , df.iloc[0, 5], df.iloc[0, 6]] 

    # Obtenemos booleano: Se añade metanol?
    adicion_metanol = df.iloc[0, 7]
    assert isinstance(adicion_metanol, (int, np.integer, float, np.floating)), \
        f"Metanol añadido debe ser numérico, no {type(adicion_metanol)}"
    adicion_metanol = False if pd.isna(adicion_metanol) or adicion_metanol == 0 else True
    # Obtiene valor metanol del ultimo mantenimiento [Cambio metanol (booleano), Cartuchos añadidos al EFOY (int), Cartuchos añadidos al STOCK (int)]
    metanol_ultimomantenimiento = [adicion_metanol , df.iloc[0, 7], df.iloc[0, 8]]
    


    # Normalizamos los valores booleanos de las columnas que contienen 'SI'/'NO'
    columnas_si_no = [2, 3, 4, 9, 10, 11, 12] # Columnas que contienen 'SI'/'NO'
    for col in columnas_si_no:
        valor = str(df.iloc[0, col]).strip().upper()  # normaliza texto
        if valor == "SI":
            df.iloc[0, col] = True
        elif valor == "NO":
            df.iloc[0, col] = False
        else:
            raise ValueError(f"Valor inesperado en la columna {col}: {valor}. Debe ser 'SI' o 'NO'.")
    
    # Comprobar si hay incidencias y generar un parte de incidencias:
    # Si alguna de las columnas de estado general del LiDAR es False, se genera un parte de incidencias.
    codigo_incidencias = []

    if df.iloc[0, 2]==False: # EFOY funciona?
        print("Advertencia: EFOY no funciona")
        codigo_incidencias.append("Error_EFOY"),
    
    if df.iloc[0, 3]==False: # Filtros de Lidar correctos?
        print("Advertencia: Filtros de Lidar no estan correctos")
        filtro_desechado = str(df.iloc[2, 3]).strip().upper()  # Filtro desechado?
        if filtro_desechado=="SI": 
            codigo_incidencias.append("Error_Filtro_Desechado")
        elif filtro_desechado=="NO":
            codigo_incidencias.append("Error_Filtro_Cambiado")
        else:
            print("Advertencia: No se ha especificado si el filtro de Lidar esta desechado o no")
            sys.exit(1)

    if df.iloc[0, 4]==False: # Escobilla limpia?
        print("Advertencia: Escobilla no esta limpia")
        codigo_incidencias.append("Error_Escobilla")

    if df.iloc[0, 9]==True: # Se cambian las baterías? Creamos una lista de estados de baterías:
        print("Advertencia: Se han cambiado las baterías")
        codigo_incidencias.append("Error_Baterias")
        estado_baterias = df.iloc[1:,9].dropna().tolist()  # Lista de estados de baterías
        # Comprobar que estado_baterias es una lista de porcentajes (int o float):
        if not all(isinstance(x, (int, float)) for x in estado_baterias):
            raise ValueError("Los estados de baterías deben ser números (int o float).")
        if not estado_baterias:
            print("Advertencia: No se ha registrado el estado de las baterías")
            estado_baterias = []
    else:
        estado_baterias = df.iloc[1:,9].dropna().tolist()  # Lista de estados de baterías
        # Comprobar que estado_baterias es una lista de porcentajes (int o float):
        if not all(isinstance(x, (int, float)) for x in estado_baterias):
            raise ValueError("Los estados de baterías deben ser números (int o float).")
        if not estado_baterias:
            print("Advertencia: No se han registrado estados de baterías")
            estado_baterias = []
        print("No se cambian las baterías, pero se han registrado sus estados.")

    if df.iloc[0, 10]==False: # Bomba de agua funciona?
        print("Advertencia: Bomba de agua no funciona")
        codigo_incidencias.append("Error_Bomba")

    if df.iloc[0, 11]==False: # Extintor de incendios revisado?
        print("Advertencia: Extintor de incendios no esta revisado")
        codigo_incidencias.append("Error_Extintor")

    if df.iloc[0, 12]==False: # Descarga de datos?
        print("Advertencia: No se han descargado los datos")
        codigo_incidencias.append("Error_DescargaDatos")
    
    if df.iloc[0, 13]: # Sensores cambiados?
        # Contar cuantas columnas de sensores cambiados hay:
        sensores_cambiados_ultimomantenimiento = df.iloc[0:, 13].dropna().tolist()
        if len(sensores_cambiados_ultimomantenimiento) > 0:
            print("Se han cambiado los siguientes sensores:", sensores_cambiados_ultimomantenimiento)
            nserie_cambio = df.iloc[0:, 14].dropna().tolist()
            nserie_recambio = df.iloc[0:, 15].dropna().tolist()
            sensores_cambiados_ultimomantenimiento = list(zip(sensores_cambiados_ultimomantenimiento, nserie_cambio, nserie_recambio))
            codigo_incidencias.append("Error_Sensores")
        else:
            sensores_cambiados_ultimomantenimiento = []

        # Si se cambia algún sensor, obtenemos el nombre del sensor cambiado y el de recambio:
    
    # Si no hay incidencias, se añade un mensaje de que todo esta correcto:
    if not codigo_incidencias:
        print("No hay incidencias, todo correcto.")

    # Si se cambia algún sensor, obtenemos el nombre del sensor cambiado y el de recambio:
   

    # Comprueba el formato de los datos leídos:
    assert isinstance(n_equipo, (str, int)), f"n_equipo debe ser str o int, no {type(n_equipo)}"
    assert pd.api.types.is_datetime64_any_dtype(df.iloc[:, 1]), f"fecha_ultimomantenimiento debe ser datetime, no {type(fecha_ultimomantenimiento)}"
    assert isinstance(comentario_ultimomantenimiento, str), f"comentario_ultimomantenimiento debe ser str, no {type(comentario_ultimomantenimiento)}"
    assert isinstance(liquido_ultimomantenimiento[0], (bool, int)), f"Cambio liquido debe ser bool o int, no {type(liquido_ultimomantenimiento[0])}"
    assert isinstance(liquido_ultimomantenimiento[1],(int, np.integer, float, np.floating)), f"Liquido añadido debe ser int o float, no {type(liquido_ultimomantenimiento[1])}"
    assert isinstance(liquido_ultimomantenimiento[2],(int, np.integer, float, np.floating)), f"Liquido restante debe ser int o float, no {type(liquido_ultimomantenimiento[2])}"
    assert isinstance(metanol_ultimomantenimiento[0], (bool, int)), f"Cambio metanol debe ser bool o int, no {type(metanol_ultimomantenimiento[0])}"
    assert isinstance(metanol_ultimomantenimiento[1],(int, np.integer, float, np.floating)), f"Cartuchos añadidos al EFOY debe ser int o float, no {type(metanol_ultimomantenimiento[1])}"
    assert isinstance(metanol_ultimomantenimiento[2],(int, np.integer, float, np.floating)), f"Cartuchos añadidos al STOCK debe ser int o float, no {type(metanol_ultimomantenimiento[2])}"
    assert isinstance(df.iloc[0, 2], bool), f"EFOY funciona debe ser SI o NO, no {type(df.iloc[0, 2])}"
    assert isinstance(df.iloc[0, 3], bool), f"Filtros de Lidar correctos debe ser SI o NO, no {type(df.iloc[0, 3])}"
    assert isinstance(df.iloc[0, 4], bool), f"Escobilla limpia debe ser SI o NO, no {type(df.iloc[0, 4])}"
    assert isinstance(df.iloc[0, 9], bool), f"Se cambian las baterías? debe ser SI o NO, no {type(df.iloc[0, 9])}"
    assert isinstance(df.iloc[0, 10], bool), f"Bomba de agua funciona debe ser SI o NO, no {type(df.iloc[0, 10])}"
    assert isinstance(df.iloc[0, 11], bool), f"Extintor de incendios revisado debe ser SI o NO, no {type(df.iloc[0, 11])}"
    assert isinstance(df.iloc[0, 12], bool), f"Descarga de datos debe ser SI o NO, no {type(df.iloc[0, 12])}"

    assert isinstance(sensores_cambiados_ultimomantenimiento, list), f"sensores_cambiados_ultimomantenimiento debe ser una lista, no {type(sensores_cambiados_ultimomantenimiento)}"



    print("Datos leídos correctamente del archivo de origen, para el número de equipo:", n_equipo)

    return n_equipo, fecha_ultimomantenimiento, comentario_ultimomantenimiento, metanol_ultimomantenimiento, liquido_ultimomantenimiento, codigo_incidencias , sensores_cambiados_ultimomantenimiento, estado_baterias






# === Helpers ===
def encontrar_fila_destino(wb, n_equipo, hoja_destino_name='Lidar Windcube'):
    """
    Localiza en la hoja destino la fila correspondiente a un equipo (n_equipo).
    Busca en la primera columna desde la fila 3 hacia abajo.
    """
    hoja_destino = wb[hoja_destino_name]
    for idx, (val,) in enumerate(hoja_destino.iter_rows(min_row=3, max_col=1, values_only=True), start=3):
        if val == n_equipo:
            return idx
    raise ValueError(f"No se encontró el equipo {n_equipo} en '{hoja_destino_name}'")


def encontrar_fila_historico(wb, n_equipo, hoja_destino_historico_name='Historico'):
    """
    Localiza en la hoja histórico la fila correspondiente a un equipo (n_equipo).
    Busca en la primera columna desde la fila 3 hacia abajo.
    """
    hoja_hist = wb[hoja_destino_historico_name]
    for idx, (val,) in enumerate(hoja_hist.iter_rows(min_row=3, max_col=1, values_only=True), start=3):
        if val == n_equipo:
            return idx
    raise ValueError(f"No se encontró el equipo {n_equipo} en '{hoja_destino_historico_name}'")


# === Funciones principales: Actulización de datos ===
def actualizar_fecha_destino(informe_destino, n_equipo, fecha_ultimomantenimiento,
                             hoja_destino_name='Lidar Windcube', hoja_destino_historico_name='Historico'):
    """
    Actualiza el archivo destino con la fecha del último mantenimiento:
    - Reemplaza la fecha en la hoja destino.
    - Añade la nueva fecha al histórico, manteniendo formato dd/mm/yyyy.
    - Si la fecha nueva coincide con la anterior, se emite advertencia y se detiene la ejecución.
    """
    # Cargar el libro de trabajo y acceder a las hojas
    wb = load_workbook(informe_destino)
    hoja_destino = wb[hoja_destino_name]
    hoja_hist = wb[hoja_destino_historico_name]

    # Buscar fila del equipo en ambas hojas
    fila_destino = encontrar_fila_destino(wb, n_equipo, hoja_destino_name)
    fila_hist = encontrar_fila_historico(wb, n_equipo, hoja_destino_historico_name)

    # Acceder a la fecha previa registrada
    fecha_penultimomantenimiento = hoja_destino.cell(row=fila_destino, column=5).value

    # Formatear fechas para comparación y visualización
    fecha_penult_str = fecha_penultimomantenimiento.strftime("%d/%m/%Y") if isinstance(fecha_penultimomantenimiento, datetime) else str(fecha_penultimomantenimiento)
    fecha_ult_str = fecha_ultimomantenimiento.strftime("%d/%m/%Y") if isinstance(fecha_ultimomantenimiento, datetime) else str(fecha_ultimomantenimiento)

    # Si las fechas coinciden, advertir y detener programa
    if fecha_penult_str == fecha_ult_str:
        print(f"Advertencia: La fecha del último mantenimiento ({fecha_ult_str}) es la misma que la del penúltimo mantenimiento. No se actualizará.")
        sys.exit(1)

    # Actualizar fecha en hoja destino
    hoja_destino.cell(row=fila_destino, column=5, value=fecha_ultimomantenimiento)

    # Actualizar histórico de fechas (concatenando si ya había valores)
    historico_fechas = hoja_hist.cell(row=fila_hist, column=2).value or ""
    historico_fechas = historico_fechas.strip()
    if historico_fechas:
        nuevo_historico_fechas = f"{historico_fechas}, \n {fecha_ult_str}"
    else:
        nuevo_historico_fechas = fecha_ult_str
    hoja_hist.cell(row=fila_hist, column=2, value=nuevo_historico_fechas)

    # Guardar y cerrar
    wb.save(informe_destino)
    wb.close()


def actualizar_comentario_destino(informe_destino, n_equipo, comentario_ultimomantenimiento,
                                  hoja_destino_name='Lidar Windcube', hoja_destino_historico_name='Historico'):
    """
    Actualiza el archivo destino con el comentario del último mantenimiento:
    - Sustituye el comentario en la hoja destino.
    - Añade el comentario previo al histórico.
    """
    wb = load_workbook(informe_destino)
    hoja_destino = wb[hoja_destino_name]
    hoja_hist = wb[hoja_destino_historico_name]

    fila_destino = encontrar_fila_destino(wb, n_equipo, hoja_destino_name)
    fila_hist = encontrar_fila_historico(wb, n_equipo, hoja_destino_historico_name)

    # Acceder al comentario previo y actualizar en hoja destino
    comentario_penult = hoja_destino.cell(row=fila_destino, column=8).value
    hoja_destino.cell(row=fila_destino, column=8, value=comentario_ultimomantenimiento)

    # Actualizar histórico con comentario previo
    historico_comentarios = hoja_hist.cell(row=fila_hist, column=3).value
    if historico_comentarios:
        nuevo_historico = f"{historico_comentarios}, \n {comentario_penult}"
    else:
        nuevo_historico = str(comentario_penult)
    hoja_hist.cell(row=fila_hist, column=3, value=nuevo_historico)

    wb.save(informe_destino)
    wb.close()


def actualizar_metanol_destino(informe_destino, n_equipo, metanol_ultimomantenimiento,
                               hoja_destino_name='Lidar Windcube', hoja_destino_historico_name='Historico'):
    """
    Actualiza las columnas del histórico relacionadas con el metanol:
    - Columna 4: metanol en stock tras el mantenimiento.
    - Columna 5: total de metanol usado por el LiDAR.
    """
    wb = load_workbook(informe_destino)
    hoja_hist = wb[hoja_destino_historico_name]

    fila_hist = encontrar_fila_historico(wb, n_equipo, hoja_destino_historico_name)

    # Desempaquetar valores de metanol [cambio, añadido EFOY, añadido stock]
    cambio_metanol, metanol_anadido_EFOY, metanol_anadido_stock = metanol_ultimomantenimiento

    if cambio_metanol:  # Si se ha cambiado metanol
        # Actualizar stock restante
        metanol_stock_restante = hoja_hist.cell(row=fila_hist, column=4).value
        hoja_hist.cell(row=fila_hist, column=4, value=metanol_stock_restante + metanol_anadido_stock - metanol_anadido_EFOY)

        # Actualizar metanol usado
        metanol_usado = hoja_hist.cell(row=fila_hist, column=5).value
        hoja_hist.cell(row=fila_hist, column=5, value=metanol_usado + metanol_anadido_EFOY)

        print("En este mantenimiento se cambió el metanol")
    else:
        print("En este mantenimiento no se cambió el metanol")

    wb.save(informe_destino)
    wb.close()


def actualizar_liquido_destino(informe_destino, n_equipo, liquido_ultimomantenimiento,
                               hoja_destino_name='Lidar Windcube', hoja_destino_historico_name='Historico'):
    """
    Actualiza las columnas del histórico relacionadas con el líquido limpiaparabrisas:
    - Columna 6: cantidad de líquido restante tras el mantenimiento.
    - Columna 7: cantidad total de líquido usado a lo largo de la historia.
    """
    wb = load_workbook(informe_destino)
    hoja_hist = wb[hoja_destino_historico_name]

    fila_hist = encontrar_fila_historico(wb, n_equipo, hoja_destino_historico_name)

    # Desempaquetar valores de líquido [adición, litros añadidos, litros restantes]
    adicion_liquido, liquido_anadido, liquido_restante = liquido_ultimomantenimiento

    if adicion_liquido:  # Si se añadió líquido
        # Actualizar cantidad restante
        hoja_hist.cell(row=fila_hist, column=6, value=liquido_restante + liquido_anadido)

        # Actualizar líquido usado en toda la historia
        liquido_usado = hoja_hist.cell(row=fila_hist, column=7).value
        hoja_hist.cell(row=fila_hist, column=7, value=liquido_usado + liquido_anadido)
        print("En este mantenimiento se añadió líquido limpiaparabrisas")
    else:  # Si no se añadió líquido
        hoja_hist.cell(row=fila_hist, column=6, value=liquido_restante)
        print("En este mantenimiento no se añadió líquido limpiaparabrisas")

    wb.save(informe_destino)
    wb.close()


def actualizar_filtros_destino(informe_destino, n_equipo, codigo_incidencias, hoja_destino_name='Lidar Windcube', hoja_destino_historico_name='Historico'):
    """
    Actualiza los filtros del archivo destino con los datos del archivo de origen.
    """
    # Cargar el libro de trabajo
    wb = load_workbook(informe_destino)
    hoja_destino_historico = wb[hoja_destino_historico_name]

    fila_hist = encontrar_fila_historico(wb, n_equipo, hoja_destino_historico_name)

    # Actualizar los contadores de filtros cambiados y desechados según el código de incidencias:
    contador_filtros_cambiados = hoja_destino_historico.cell(row=fila_hist, column=8).value or 0
    if "Error_Filtro_Cambiado" in codigo_incidencias:
        hoja_destino_historico.cell(row=fila_hist, column=8, value=contador_filtros_cambiados + 1)
        # Actualizar la columna 9 con el contador de filtros desechados
        contador_filtros_desechados = hoja_destino_historico.cell(row=fila_hist, column=9).value or 0
    if "Error_Filtro_Desechado" in codigo_incidencias:
        hoja_destino_historico.cell(row=fila_hist, column=9, value=contador_filtros_desechados + 1)
            
    # Guardar los cambios en el archivo destino
    wb.save(informe_destino)
    wb.close()



def actualizar_escobilla_destino(informe_destino, n_equipo, codigo_incidencias, hoja_destino_name='Lidar Windcube', hoja_destino_historico_name='Historico'):
    """
    Actualiza los filtros del archivo destino con los datos del archivo de origen.
    """
    # Cargar el libro de trabajo
    wb = load_workbook(informe_destino)
    hoja_destino = wb[hoja_destino_name]
    hoja_destino_historico = wb[hoja_destino_historico_name]

    fila_hist = encontrar_fila_historico(wb, n_equipo, hoja_destino_historico_name)

    # Actualizar la columna 10 con el contador de escobillas cambiadas
    contador_escobillas_cambiadas = hoja_destino_historico.cell(row=fila_hist, column=10).value or 0
    if "Error_Escobilla" in codigo_incidencias:
        hoja_destino_historico.cell(row=fila_hist, column=10, value=contador_escobillas_cambiadas + 1)
    # Guardar los cambios en el archivo destino
    wb.save(informe_destino)
    wb.close()



def actualizar_incidencias_destino(informe_destino, n_equipo, codigo_incidencias,
                                   hoja_destino_name='Lidar Windcube', hoja_destino_historico_name='Historico'):
    """
    Actualiza el archivo destino añadiendo incidencias al comentario:
    - Se concatenan al comentario de la hoja destino.
    - Se concatenan al comentario histórico.
    """
    wb = load_workbook(informe_destino)
    hoja_destino = wb[hoja_destino_name]
    hoja_hist = wb[hoja_destino_historico_name]

    fila_destino = encontrar_fila_destino(wb, n_equipo, hoja_destino_name)
    fila_hist = encontrar_fila_historico(wb, n_equipo, hoja_destino_historico_name)

    # Obtener comentarios actuales (último y histórico)
    comentario_ult = hoja_destino.cell(row=fila_destino, column=8).value or ""
    comentario_hist = hoja_hist.cell(row=fila_hist, column=3).value or ""

    # Si hay incidencias, concatenarlas
    if codigo_incidencias:
        incidencias_str = ", ".join(codigo_incidencias)
        if comentario_ult:
            comentario_ult += f" | Incidencias: {incidencias_str}"
        else:
            comentario_ult = f"Incidencias: {incidencias_str}"

        if comentario_hist:
            comentario_hist += f" | Incidencias: {incidencias_str}"
        else:
            comentario_hist = f"Incidencias: {incidencias_str}"

    # Guardar comentarios actualizados
    hoja_destino.cell(row=fila_destino, column=8, value=comentario_ult)
    hoja_hist.cell(row=fila_hist, column=3, value=comentario_hist)

    wb.save(informe_destino)
    wb.close()



def actualizar_baterias_destino(informe_destino, n_equipo, codigo_incidencias, estado_baterias,
                                hoja_destino_name='Lidar Windcube', hoja_destino_historico_name='Historico'):
    """
    Actualiza el estado de las baterías en el archivo destino:
    - Si todas están en buen estado, escribe la proporción de baterías en buen estado.
    - Si hay baterías en mal estado, añade una incidencia y actualiza el histórico con los detalles.
    - Si los valores están fuera de rango (0-100), lanza un error.
    """

    # Cargar el libro de trabajo
    wb = load_workbook(informe_destino)
    hoja_destino = wb[hoja_destino_name]
    hoja_hist = wb[hoja_destino_historico_name]

    # Buscar fila del equipo tanto en hoja destino como en histórico
    fila_destino = encontrar_fila_destino(wb, n_equipo, hoja_destino_name)
    fila_hist = encontrar_fila_historico(wb, n_equipo, hoja_destino_historico_name)

    # Contadores para clasificar el estado de las baterías
    baterias_buen_estado = 0
    baterias_mal_estado = 0

    # Clasificar cada batería según su SOH (State of Health)
    for soh in estado_baterias:
        if 80 <= soh <= 100:
            baterias_buen_estado += 1
        elif 0 <= soh < 80:
            baterias_mal_estado += 1
        else:
            raise ValueError(f"Estado de batería {soh} fuera de rango (0-100).")

    # Si no hay datos de baterías, emitir advertencia y salir
    if baterias_buen_estado == 0 and baterias_mal_estado == 0:
        print("Advertencia: No se ha registrado el estado de las baterías")
        wb.close()
        return

    # Caso 1: Hay baterías en mal estado
    if baterias_mal_estado > 0:
        # Obtener valor previo en el histórico (columna 11)
        estados_baterias_pasado = hoja_hist.cell(row=fila_hist, column=11).value

        if "Error_Baterias" in codigo_incidencias:
            # Registrar que hay baterías en mal estado y se deben cambiar
            nuevo_estado = f"{baterias_mal_estado} de {baterias_buen_estado + baterias_mal_estado} baterías en mal estado. Se deben cambiar"
            print("Advertencia: Baterías no están en buen estado. Se deben cambiar")
        else:
            # Registrar que hubo baterías en mal estado pero ya se cambiaron
            nuevo_estado = f"{baterias_mal_estado} de {baterias_buen_estado + baterias_mal_estado} baterías en mal estado. Se han cambiado"
            print("Advertencia: Baterías no estaban en buen estado. Se han cambiado")

        # Concatenar al histórico si ya había valores
        if estados_baterias_pasado:
            hoja_hist.cell(row=fila_hist, column=11, value=f"{estados_baterias_pasado}, \n {nuevo_estado}")
        else:
            hoja_hist.cell(row=fila_hist, column=11, value=nuevo_estado)

    # Caso 2: Todas las baterías están en buen estado
    elif baterias_buen_estado > 0 and baterias_mal_estado == 0:
        hoja_hist.cell(row=fila_hist, column=11,
                       value=f"{baterias_buen_estado} de {baterias_buen_estado + baterias_mal_estado} baterías en buen estado")
        print("Todas las baterías están en buen estado")

    # Guardar cambios en el archivo destino
    wb.save(informe_destino)
    wb.close()



def actualizar_sensores_destino(informe_destino, n_equipo, sensores_cambiados_ultimomantenimiento,
                                 hoja_destino_name='Lidar Windcube', hoja_destino_historico_name='Historico'):
    """
    Actualiza los sensores en el archivo destino:
    - Añade los sensores cambiados al historico (columna 12), en formato "Sensor cambiado - (Nº serie antiguo) - (Nº serie recambio)"
    - Añade un string append al final del comentario del último mantenimiento en la hoja destino y el historico advertiendo que se cambio el sensor x.
    """

    # Cargar el libro de trabajo
    wb = load_workbook(informe_destino)
    hoja_destino = wb[hoja_destino_name]
    hoja_hist = wb[hoja_destino_historico_name]
    # Buscar fila del equipo tanto en hoja destino como en histórico
    fila_destino = encontrar_fila_destino(wb, n_equipo, hoja_destino_name)
    fila_hist = encontrar_fila_historico(wb, n_equipo, hoja_destino_historico_name) 

    # Si no se cambiaron sensores, salir
    if not sensores_cambiados_ultimomantenimiento:
        print("No se han cambiado sensores en este mantenimiento")
        wb.close()
        return
    
    # Si se cambiaron sensores, actualizar histórico
    sensores_cambiados_str = []
    for sensor, nserie_cambio, nserie_recambio in sensores_cambiados_ultimomantenimiento:
        sensores_cambiados_str.append(f"{sensor} - ({nserie_cambio}) - ({nserie_recambio})")
    sensores_cambiados_str = ", ".join(sensores_cambiados_str)  
    # Actualizar columna 12 del histórico
    sensores_previos = hoja_hist.cell(row=fila_hist, column=12).value or ""
    if sensores_previos:
        hoja_hist.cell(row=fila_hist, column=12, value=f"{sensores_previos}, \n {sensores_cambiados_str}")
    else:
        hoja_hist.cell(row=fila_hist, column=12, value=sensores_cambiados_str)
    # Actualizar comentario en hoja destino
    comentario_ult = hoja_destino.cell(row=fila_destino, column=8).value or ""
    if comentario_ult:
        hoja_destino.cell(row=fila_destino, column=8, value=f"{comentario_ult} | Sensores cambiados: {sensores_cambiados_str}")
    else:
        hoja_destino.cell(row=fila_destino, column=8, value=f"Sensores cambiados: {sensores_cambiados_str}")

    # Guardar cambios en el archivo destino
    wb.save(informe_destino)
    wb.close()  
