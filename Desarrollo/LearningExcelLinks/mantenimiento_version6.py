# mantenimiento.py

import pandas as pd
import numpy as np
import sys
from openpyxl import load_workbook
from datetime import datetime

# Leer archivo de origen con la siguiente función:

def leer_datos_origen(informe_mtto, hoja_mtto):
    """
    Lee los datos del archivo de origen, y devuelve un DataFrame si y solo si los datos tienen el formsto correcto.
    """
    df = pd.read_excel(informe_mtto, sheet_name=hoja_mtto, engine='openpyxl')

    # Convertir "Cambio Metanol: SI/NO" a booleano:
    if df.iloc[0, 6] == 'SI':
        df.iloc[0, 6] = True
    elif df.iloc[0, 6] == 'NO':
        df.iloc[0, 6] = False
    ## valor = df.iloc[0, 6]  # columna 7
    ##df.iloc[0, 6] = bool(valor) and valor != 0

    
    # Convertir  "Cambio Líquido limpiapara: SI/NO" a booleano:
    if df.iloc[0, 3] == 'SI':
        df.iloc[0, 3] = True
    elif df.iloc[0, 3] == 'NO':
        df.iloc[0, 3] = False

    n_equipo = df.iloc[0, 0]  # Obtiene valor n equipo LiDAR
    fecha_ultimomantenimiento = df.iloc[0, 1]  # Obtiene valor fecha del ultimo mantenimiento
    comentario_ultimomantenimiento = df.iloc[0, 15]  # Obtiene valor comentario del ultimo mantenimiento
    # Obtiene valor metanol del ultimo mantenimiento [Cambio metanol (booleano), Cartuchos añadidos al EFOY (int), Cartuchos añadidos al STOCK (int)]
    metanol_ultimomantenimiento = [df.iloc[0, 6] , df.iloc[0, 7], df.iloc[0, 8]]

    # Obtiene valor liquido parabrisas del ultimo mantenimiento [Cambio liquido (booleano), Liquido añadido (int), Liquido restante (int)]  
    liquido_ultimomantenimiento = [df.iloc[0, 3] , df.iloc[0, 4], df.iloc[0, 5]] 



    # Comprueba el formato de los datos leídos:
    assert isinstance(n_equipo, (str, int)), f"n_equipo debe ser str o int, no {type(n_equipo)}"
    assert pd.api.types.is_datetime64_any_dtype(df.iloc[:, 1]), f"fecha_ultimomantenimiento debe ser datetime, no {type(fecha_ultimomantenimiento)}"
    assert isinstance(comentario_ultimomantenimiento, str), f"comentario_ultimomantenimiento debe ser str, no {type(comentario_ultimomantenimiento)}"
    assert isinstance(metanol_ultimomantenimiento[0], (bool, int)), f"Cambio metanol debe ser bool o int, no {type(metanol_ultimomantenimiento[0])}"
    assert isinstance(metanol_ultimomantenimiento[1],(int, np.integer)), f"Cartuchos añadidos al EFOY debe ser int o float, no {type(metanol_ultimomantenimiento[1])}"
    assert isinstance(metanol_ultimomantenimiento[2],(int, np.integer)), f"Cartuchos añadidos al STOCK debe ser int o float, no {type(metanol_ultimomantenimiento[2])}"
    assert isinstance(liquido_ultimomantenimiento[0], (bool, int)), f"Cambio liquido debe ser bool o int, no {type(liquido_ultimomantenimiento[0])}"
    assert isinstance(liquido_ultimomantenimiento[1],(int, np.integer, float, np.floating)), f"Liquido añadido debe ser int o float, no {type(liquido_ultimomantenimiento[1])}"
    assert isinstance(liquido_ultimomantenimiento[2],(int, np.integer, float, np.floating)), f"Liquido restante debe ser int o float, no {type(liquido_ultimomantenimiento[2])}"


    # assert isinstance(metanol_ultimomantenimiento[2],(int, float, np.floating)), f"Cartuchos añadidos al STOCK debe ser int o float, no {type(metanol_ultimomantenimiento[1])}"


    print("Datos leídos correctamente del archivo de origen, para el número de equipo:", n_equipo)

    return n_equipo, fecha_ultimomantenimiento, comentario_ultimomantenimiento, metanol_ultimomantenimiento, liquido_ultimomantenimiento






#Actualizar el archivo destino con los datos leídos del archivo de origen, a partir de las siguientes funciones:

def actualizar_fecha_destino(informe_destino, n_equipo, fecha_ultimomantenimiento, hoja_destino_name='Lidar Windcube', hoja_destino_historico_name='Historico'):
    """
    Actualiza el archivo destino con la fecha del archivo de origen.
    """
    # Cargar el libro de trabajo
    wb = load_workbook(informe_destino)
    hoja_destino = wb[hoja_destino_name]
    hoja_destino_historico = wb[hoja_destino_historico_name]

    # Buscar la fila donde la primera columna coincide con n_equipo LiDAR,
    # copiar la fecha de ultimo mantenimiento desde el archivo de origen y hacer append en la columna 2 de la hoja 'Historico',
    # y actualizar la columna 5 con la nueva fecha de ultimo mantenimiento:
    for idx, row in enumerate(hoja_destino.iter_rows(min_row=3, max_col=1, values_only=True), start=3):
        if row[0] == n_equipo:
            # Accede fecha del penultimo mantenimiento, presente en el excel padre
            fecha_penultimomantenimiento = hoja_destino.cell(row=idx, column=5).value

            # Formatear fechas para que siempre sean dd/mm/yyyy
            fecha_penult_str = fecha_penultimomantenimiento.strftime("%d/%m/%Y") if isinstance(fecha_penultimomantenimiento, datetime) else str(fecha_penultimomantenimiento)
            fecha_ult_str   = fecha_ultimomantenimiento.strftime("%d/%m/%Y") if isinstance(fecha_ultimomantenimiento, datetime) else str(fecha_ultimomantenimiento)
            if fecha_penult_str == fecha_ult_str:
                print(f"Advertencia: La fecha del último mantenimiento ({fecha_ult_str}) es la misma que la del penúltimo mantenimiento. No se actualizará.")
                sys.exit(1)
            # Actualizar la columna 5 con la fecha del último mantenimiento
            hoja_destino.cell(row=idx, column=5, value=fecha_ultimomantenimiento)
            # Concatenar la fecha del último mantenimiento en la hoja 'Historico'
            historico_fechas = hoja_destino_historico.cell(row=idx, column=2).value 

            # Esta función toma un valor que puede ser una fecha, una cadena de fechas separadas   
            # por comas, o None, y devuelve una cadena con las fechas formateadas.
            def formatear_varias_fechas(valor):
                if valor is None:
                    return ""
                if isinstance(valor, datetime):
                    return valor.strftime("%d/%m/%Y")
                if isinstance(valor, str):
                    partes = [p.strip() for p in valor.split(",")]
                    nuevas_partes = []
                    for parte in partes:
                        try:
                            fecha = datetime.strptime(parte, "%Y-%m-%d %H:%M:%S")
                            nuevas_partes.append(fecha.strftime("%d/%m/%Y"))
                        except ValueError:
                            nuevas_partes.append(parte)  # No era una fecha con ese formato
                    return ", \n ".join(nuevas_partes) # Une elementos de una lista de strings en una sola string separada por comas y enters
                return str(valor)
            
            # Formatear las fechas del historico
            historico_fechas = formatear_varias_fechas(historico_fechas)

            if historico_fechas:
                # Si ya hay un valor, concatenar como una cadena separada por comas
                nuevo_historico_fechas = f"{historico_fechas}, \n {fecha_ult_str}"
            else:
                # Sino, iniciar con la fecha del último mantenimiento
                nuevo_historico_fechas = str(fecha_ult_str)

            hoja_destino_historico.cell(row=idx, column=2, value=nuevo_historico_fechas)
            break
    # Guardar los cambios en el archivo destino
    wb.save(informe_destino)
    wb.close()




def actualizar_comentario_destino(informe_destino, n_equipo, comentario_ultimomantenimiento, hoja_destino_name='Lidar Windcube', hoja_destino_historico_name='Historico'):
    """
    Actualiza el archivo destino con el comentario del archivo de origen.
    """
    # Cargar el libro de trabajo
    wb = load_workbook(informe_destino)
    hoja_destino = wb[hoja_destino_name]
    hoja_destino_historico = wb[hoja_destino_historico_name]

    # Buscar la fila donde la primera columna coincide con n_equipo LiDAR,
    # copiar el comentario del ultimo mantenimiento desde el archivo de origen, hacer append en la columna 3 de la hoja 'Historico',
    # y actualizar la columna 8 con el comentario del ultimo mantenimiento:
    for idx, row in enumerate(hoja_destino.iter_rows(min_row=3, max_col=1, values_only=True), start=3):
        if row[0] == n_equipo:
            # Accede comentario del penultimo mantenimiento, presente en el excel padre
            comentario_penultimomantenimiento = hoja_destino.cell(row=idx, column=8).value
            # Actualizar la columna 8 con el comentario del ultimo mantenimiento
            hoja_destino.cell(row=idx, column=8, value=comentario_ultimomantenimiento)
            # Concatenar el comentario del último mantenimiento en la hoja 'Historico'
            historico_comentarios = hoja_destino_historico.cell(row=idx, column=3).value 
            if historico_comentarios:
                # Si ya hay un valor, concatenar como una cadena separada por comas
                nuevo_historico_comentarios = f"{historico_comentarios}, \n {comentario_ultimomantenimiento}"
            else:
                # Sino, iniciar con el comentario del último mantenimiento
                nuevo_historico_comentarios = str(comentario_ultimomantenimiento)

            hoja_destino_historico.cell(row=idx, column=3, value=nuevo_historico_comentarios)
            break
    # Guardar los cambios en el archivo destino
    wb.save(informe_destino)
    wb.close()




def actualizar_metanol_destino(informe_destino, n_equipo, metanol_ultimomantenimiento, hoja_destino_name='Lidar Windcube', hoja_destino_historico_name='Historico'):
    """
    Actualiza las columnas 4 y 5 del archivo destino con el metanol del archivo de origen.
    """
    # Cargar el libro de trabajo
    wb = load_workbook(informe_destino)
    hoja_destino = wb[hoja_destino_name]
    hoja_destino_historico = wb[hoja_destino_historico_name]
    cambio_metanol = metanol_ultimomantenimiento[0]  # Booleano que indica si se ha cambiado el metanol
    metanol_anadido_EFOY = metanol_ultimomantenimiento[1]  # Cantidad de metanol añadido al EFOY
    metanol_anadido_stock = metanol_ultimomantenimiento[2]  # Cantidad de metanol añadido al stock



    # Buscar la fila donde la primera columna coincide con n_equipo LiDAR,
    # Y actualizar el metanol si y solo si se ha cambiado:
    if cambio_metanol:  # Si se ha cambiado el metanol
        for idx, row in enumerate(hoja_destino.iter_rows(min_row=3, max_col=1, values_only=True), start=3):
            if row[0] == n_equipo:
                # Actualizar la columna 4 con el valor de cartuchos de metanol en STOCK tras el mantenimiento
                metanol_stock_restante = hoja_destino_historico.cell(row=idx, column=4).value # Metanol en stock antes del mantenimiento
                hoja_destino_historico.cell(row=idx, column=4, value= metanol_stock_restante + metanol_anadido_stock - metanol_anadido_EFOY )  # Actualizamos el metanol en stock despues el mantenimiento
                # Actualizar la columna 5 con el valor de cartuchos de metanol gastados en la historia del LiDAR (= Cartuchos totales instalados en el EFOY)
                metanol_usado = hoja_destino_historico.cell(row=idx, column=5).value # Metanol usado por el LiDAR hasta el mantenimiento anterior
                hoja_destino_historico.cell(row=idx, column=5, value= metanol_usado + metanol_anadido_EFOY )  # Actualizamos el metanol usado por el EFOY, despues el mantenimiento
                print("En este mantenimiento se cambio el metanol")
                break

    elif cambio_metanol == False:
        print("En este mantenimiento no se cambio el metanol")
                

    # Guardar los cambios en el archivo destino
    wb.save(informe_destino)
    wb.close()



def actualizar_liquido_destino(informe_destino, n_equipo, liquido_ultimomantenimiento, hoja_destino_name='Lidar Windcube', hoja_destino_historico_name='Historico'):
    """
    Actualiza las columnas 6 y 7 del archivo destino con el liquido del archivo de origen.
    """
    # Cargar el libro de trabajo
    wb = load_workbook(informe_destino)
    hoja_destino = wb[hoja_destino_name]
    hoja_destino_historico = wb[hoja_destino_historico_name]
    adicion_liquido = liquido_ultimomantenimiento[0]  # Booleano que indica si se ha cambiado el liquido
    liquido_anadido = liquido_ultimomantenimiento[1]  # Cantidad de liquido añadido al deposito de liquido limpiaparabrisas
    liquido_restante = liquido_ultimomantenimiento[2]  # Cantidad de liquido restante en el deposito de liquido limpiaparabrisas al llegar al mantenimiento



    # Buscar la fila donde la primera columna coincide con n_equipo LiDAR,
    # Y actualizar el liquido si y solo si se ha cambiado:
    if adicion_liquido:  # Si se ha cambiado el liquido
        for idx, row in enumerate(hoja_destino.iter_rows(min_row=3, max_col=1, values_only=True), start=3):
            if row[0] == n_equipo:
                # Actualizar la columna 6 con el valor de liquido restante tras el mantenimiento
                # liquido_restante_penultimomantenimineto = hoja_destino_historico.cell(row=idx, column=6).value # Liquido restante antes al llegar en el penulmantenimiento
                hoja_destino_historico.cell(row=idx, column=6, value= liquido_restante + liquido_anadido)  # Actualizamos el liquido restante despues el mantenimiento
                # Actualizar la columna 7 con el valor de litros de liquido gastados en la historia del LiDAR (= Litros totales echados en el deposito de liquido limpiaparabrisas)
                liquido_usado = hoja_destino_historico.cell(row=idx, column=7).value # Liquido usado por el LiDAR hasta el mantenimiento anterior
                hoja_destino_historico.cell(row=idx, column=7, value = liquido_usado + liquido_anadido)  # Actualizamos el liquido usado por el limpiaparabrisas, despues el mantenimiento
                # !!!!!!!!!!!!!!!!!!!
                # Podría cambiar el liquido usado a: value = liquido_usado + (liquido_restante_penultimomantenimineto - (liquido_restante+liquido_anadido))
                print("En este mantenimiento se añadio liquido limpiaparabrisas")
                break

    elif adicion_liquido == False:
        for idx, row in enumerate(hoja_destino.iter_rows(min_row=3, max_col=1, values_only=True), start=3):
            if row[0] == n_equipo:
                # Actualizar la columna 6 con el valor de liquido restante tras el mantenimiento
                # liquido_restante_penultimomantenimineto = hoja_destino_historico.cell(row=idx, column=6).value # Liquido restante antes del mantenimiento, al llegar
                hoja_destino_historico.cell(row=idx, column=6, value= liquido_restante)  # Actualizamos el liquido restante despues el mantenimiento
                print("En este mantenimiento no se añadio liquido limpiaparabrisas")
                break

    # Guardar los cambios en el archivo destino
    wb.save(informe_destino)
    wb.close()






def actualizar

