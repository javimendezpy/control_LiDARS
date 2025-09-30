import pandas as pd
# Importar la librería pandas para manejar DataFrames
# Importar la librería openpyxl para trabajar con archivos Excel
from openpyxl import load_workbook



# Definimos nombre del archivo de origen y destino
informe_mtto = r'\\S31889004.services.dekra.com\DKI\EOLICA\MAT REF\BASE DE DATOS ESTACIONES\LearningExcelLinks\informe_mtto.xlsx'
informe_destino = r'\\S31889004.services.dekra.com\DKI\EOLICA\MAT REF\BASE DE DATOS ESTACIONES\LearningExcelLinks\ExcelPadre.xlsx'
# Definimos nombre de las hojas de origen y destino
hoja_mtto = 'Hoja1' #Nombre de la hoja del archivo de origen
hoja_destino = 'Lidar Windcube' #Nombre de la hoja del archivo destino
hoja_destino_historico = 'Historico' #Nombre de la hoja del archivo destino para el historico



# Leer archivo de origen con la siguiente función:

def leer_datos_origen(informe_mtto, hoja_mtto):
    """
    Lee los datos del archivo de origen y devuelve un DataFrame.
    """
    df = pd.read_excel(informe_mtto, sheet_name=hoja_mtto, engine='openpyxl')
    n_equipo = df.iloc[0, 0]  # Obtiene valor n equipo LiDAR
    fecha_ultimomantenimiento = df.iloc[0, 1]  # Obtiene valor fecha del ultimo mantenimiento
    comentario_ultimomantenimiento = df.iloc[0, 12]  # Obtiene valor comentario del ultimo mantenimiento
    return n_equipo, fecha_ultimomantenimiento, comentario_ultimomantenimiento




#Actualizar el archivo destino con los datos leídos del archivo de origen, a partir de las siguientes funciones:

def actualizar_fecha_destino(informe_destino, n_equipo, fecha_ultimomantenimiento, hoja_destino_name='Lidar Windcube', hoja_destino_historico_name='Historico'):
    """
    Actualiza el archivo destino con la fecha del archivo de origen.
    """
    # Cargar el libro de trabajo
    wb = load_workbook(informe_destino)
    hoja_destino = wb[hoja_destino_name]
    hoja_destino_historico = wb[hoja_destino_historico_name]

    # Buscar la fila donde la primera columna coincide con n_equipo LiDAR,
    # copiar la fecha de ultimo mantenimiento desde el archivo de origen y hacer append en la columna 2 de la hoja 'Historico',
    # y actualizar la columna 5 con la nueva fecha de ultimo mantenimiento:
    for idx, row in enumerate(hoja_destino.iter_rows(min_row=3, max_col=1, values_only=True), start=3):
        if row[0] == n_equipo:
            # Accede fecha del penultimo mantenimiento, presente en el excel padre
            fecha_penultimomantenimiento = hoja_destino.cell(row=idx, column=5).value
            # Actualizar la columna 5 con la fecha del último mantenimiento
            hoja_destino.cell(row=idx, column=5, value=fecha_ultimomantenimiento)
            # Concatenar la fecha del penúltimo mantenimiento en la hoja 'Historico'
            historico_fechas = hoja_destino_historico.cell(row=idx, column=2).value 
            if historico_fechas:
                # Si ya hay un valor, concatenar como una cadena separada por comas
                nuevo_historico_fechas = f"{historico_fechas}, \n {fecha_penultimomantenimiento}"
            else:
                # Sino, iniciar con la fecha del penúltimo mantenimiento
                nuevo_historico_fechas = str(fecha_penultimomantenimiento)

            hoja_destino_historico.cell(row=idx, column=2, value=nuevo_historico_fechas)
            break
    # Guardar los cambios en el archivo destino
    wb.save(informe_destino)
    wb.close()




def actualizar_comentario_destino(informe_destino, n_equipo, comentario_ultimomantenimiento, hoja_destino_name='Lidar Windcube', hoja_destino_historico_name='Historico'):
    """
    Actualiza el archivo destino con el comentario del archivo de origen.
    """
    # Cargar el libro de trabajo
    wb = load_workbook(informe_destino)
    hoja_destino = wb[hoja_destino_name]
    hoja_destino_historico = wb[hoja_destino_historico_name]

    # Buscar la fila donde la primera columna coincide con n_equipo LiDAR,
    # copiar el comentario del ultimo mantenimiento desde el archivo de origen y hacer append en la columna 3 de la hoja 'Historico',
    # y actualizar la columna 8 con el comentario del ultimo mantenimiento:
    for idx, row in enumerate(hoja_destino.iter_rows(min_row=3, max_col=1, values_only=True), start=3):
        if row[0] == n_equipo:
            # Accede comentario del penultimo mantenimiento, presente en el excel padre
            comentario_penultimomantenimiento = hoja_destino.cell(row=idx, column=8).value
            # Actualizar la columna 8 con el comentario del ultimo mantenimiento
            hoja_destino.cell(row=idx, column=8, value=comentario_ultimomantenimiento)
            # Concatenar el comentario del penúltimo mantenimiento en la hoja 'Historico'
            historico_comentarios = hoja_destino_historico.cell(row=idx, column=3).value 
            if historico_comentarios:
                # Si ya hay un valor, concatenar como una cadena separada por comas
                nuevo_historico_comentarios = f"{historico_comentarios}, \n {comentario_penultimomantenimiento}"
            else:
                # Sino, iniciar con el comentario del penúltimo mantenimiento
                nuevo_historico_comentarios = str(comentario_penultimomantenimiento)

            hoja_destino_historico.cell(row=idx, column=3, value=nuevo_historico_comentarios)
            break
    # Guardar los cambios en el archivo destino
    wb.save(informe_destino)
    wb.close()




def main():
    # Leer datos del archivo de origen
    n_equipo, fecha_ultimomantenimiento, comentario_ultimomantenimiento = leer_datos_origen(informe_mtto, hoja_mtto)
    
    # Actualizar el archivo destino con los datos leídos
    actualizar_fecha_destino(informe_destino, n_equipo, fecha_ultimomantenimiento, hoja_destino, hoja_destino_historico)
    # Actualizar el comentario del archivo destino con los datos leídos
    actualizar_comentario_destino(informe_destino, n_equipo, comentario_ultimomantenimiento, hoja_destino, hoja_destino_historico)  



# Ejecutar la función principal
# Esta parte del código se ejecuta cuando el script se corre directamente
# Si se importa como un módulo, no se ejecutará automáticamente, ya que main() no se llamará.
if __name__ == "__main__":
    main()
    print("Datos actualizados correctamente.")
    # Este script actualiza el archivo destino con los datos del archivo de origen.
    # Asegúrate de que las rutas de los archivos y los nombres de las hojas sean correctos.